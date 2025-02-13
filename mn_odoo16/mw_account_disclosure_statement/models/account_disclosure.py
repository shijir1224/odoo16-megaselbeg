# -*- encoding: utf-8 -*-
##############################################################################
from odoo import api, fields, models, _
from odoo.modules import get_module_resource
import openpyxl
import time
import base64
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

TYPE_SELECTION = [('1', u'Батлагдсан гүйлгээ'),
                  ('2', u'Бүх гүйлгээ')]
"""
   Санхүүгийн тайлангийн тодруулга тайлан
"""


class AccountDisclosureMain(models.Model):
    _name = "account.disclosure.main"
    _description = u"Санхүүгийн тайлангийн тодруулга үндсэн тайлан"

    name = fields.Char(u'Нэр', required=True)
    company_id = fields.Many2one('res.company', u'Компани', required=True)
    type_n = fields.Selection(TYPE_SELECTION, u'Гүйлгээний шүүлт', default='1', required=True)
    fiscalyear_id = fields.Many2one('account.fiscalyear', u'Санхүүгийн жил', required=True)

    intro1 = fields.Text(u'1. ТАЙЛАН БЭЛТГЭХ ҮНДЭСЛЭЛ')
    intro2 = fields.Text(u'2. НЯГТЛАН БОДОХ БҮРТГЭЛИЙН БОДЛОГЫН ӨӨРЧЛӨЛТ')

    bolzoshgui_horongo_ba_or_tolbor = fields.Text(u'23. БОЛЗОШГҮЙ ХӨРӨНГӨ БА ӨР ТӨЛБӨР')
    tailagnaliin_uyiin_daraah_uil_yavts = fields.Text(u'24. ТАЙЛАГНАЛЫН ҮЕИЙН ДАРААХ ҮЙЛ ЯВЦ')
    # 17.4 Эздийн өмчийн бусад хэсэг
    ezdiin_omchiin_busad_heseg = fields.Text(u'17.4 Эздийн өмчийн бусад хэсэг')

    page1_1_ids = fields.One2many('account.disclosure.main.page1_1', 'main_id',
                                  u'Үндсэн үйл ажиллагааны чиглэл /төрөл/')
    page1_2_ids = fields.One2many('account.disclosure.main.page1_2', 'main_id',
                                  u'Туслах үйл ажиллагааны чиглэл /төрөл/')
    page1_3_ids = fields.One2many('account.disclosure.main.page1_3', 'main_id',
                                  u'Салбар, төлөөлөгчийн газрын нэр, байршил')
    page3_ids = fields.One2many('account.disclosure.main.page3', 'main_id', u'3. МӨНГӨ БА ТҮҮНТЭЙ АДИЛТГАХ ХӨРӨНГӨ')

    page4_1_ids = fields.One2many('account.disclosure.main.page4_1', 'main_id', u'4.1 Дансны авлага')
    page4_2_ids = fields.One2many('account.disclosure.main.page4_2', 'main_id',
                                  u'4.2 Татвар, нийгмийн даатгалын шимтгэл (НДШ)-ийн  авлага')
    page4_3_ids = fields.One2many('account.disclosure.main.page4_3', 'main_id',
                                  u'4.3  Бусад богино хугацаат авлага (төрлөөр ангилна)')
    page5_ids = fields.One2many('account.disclosure.main.page5', 'main_id', u'5. БУСАД САНХҮҮГИЙН ХӨРӨНГӨ')
    page6_ids = fields.One2many('account.disclosure.main.page6', 'main_id', u'6. БАРАА МАТЕРИАЛ')
    page17_3_ids = fields.One2many('account.disclosure.main.page17_3', 'main_id',
                                   u'17.3 Гадаад валютын хөрвүүлэлтийн нөөц')
    page18_1_ids = fields.One2many('account.disclosure.main.page18_1', 'main_id',
                                   u'18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ')
    page18_2_ids = fields.One2many('account.disclosure.main.page18_2', 'main_id',
                                   u'18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ')
    page18_3_ids = fields.One2many('account.disclosure.main.page18_3', 'main_id',
                                   u'18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ')
    page18_4_ids = fields.One2many('account.disclosure.main.page18_4', 'main_id',
                                   u'18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ')
    page19_1_ids = fields.One2many('account.disclosure.main.page19_1', 'main_id',
                                   u'19.1 Бусад орлого')
    page19_2_ids = fields.One2many('account.disclosure.main.page19_2', 'main_id',
                                   u'19.2 Гадаад валютын ханшийн зөрүүний олз, гарз')
    page19_3_ids = fields.One2many('account.disclosure.main.page19_3', 'main_id', u'19.3 Бусад ашиг (алдагдал)')
    page20_1_ids = fields.One2many('account.disclosure.main.page20_1', 'main_id',
                                   u'20.1 Борлуулалт маркетингийн болон ерөнхий ба удирдлагын зардлууд')
    page20_2_ids = fields.One2many('account.disclosure.main.page20_2', 'main_id', u'20.2 Бусад зардал')
    page20_3_ids = fields.One2many('account.disclosure.main.page20_3', 'main_id', u'20.3 Цалингийн зардал')
    page21_ids = fields.One2many('account.disclosure.main.page21', 'main_id',
                                 u'21. ОРЛОГЫН ТАТВАРЫН ЗАРДАЛ')
    page22_1_ids = fields.One2many('account.disclosure.main.page22_1', 'main_id',
                                   u'22.1 Толгой компани, хамгийн дээд хяналт тавигч '
                                   u'компани, хувь хүний талаарх мэдээлэл****')
    page22_2_ids = fields.One2many('account.disclosure.main.page22_2', 'main_id',
                                   u'22.2 Тэргүүлэх удирдлагын бүрэлдэхүүнд олгосон нөхөн олговрын тухай мэдээлэл')
    page22_3_ids = fields.One2many('account.disclosure.main.page22_3', 'main_id',
                                   u'22.3 Холбоотой талуудтай хийсэн ажил гүйлгээ')
    page8_ids = fields.One2many('account.disclosure.main.page8', 'main_id', u'8. УРЬДЧИЛЖ ТӨЛСӨН ЗАРДАЛ/ТООЦОО')
    page9_1_ids = fields.One2many('account.disclosure.main.page9_1', 'main_id', u'9. ҮНДСЭН ХӨРӨНГӨ ')
    page9_2_ids = fields.One2many('account.disclosure.main.page9_2', 'main_id', u'9. ҮНДСЭН ХӨРӨНГӨ ')
    page10_1_ids = fields.One2many('account.disclosure.main.page10', 'main_id', u'10. БИЕТ БУС ХӨРӨНГӨ ХӨРӨНГӨ ')
    page4_describe = fields.Text(u'4 - Тэмдэглэл. (Дансны авлагыг төлөгдөх хугацаандаа байгаа, хугацаа хэтэрсэн, '
                                 u'төлөгдөх найдваргүй гэж ангилна. Найдваргүй авлагын хасагдуулга байгуулсан арга, '
                                 u'гадаад валютаар илэрхийлэгдсэн авлагын талаар болон бусад тайлбар '
                                 u'тэмдэглэлийг хийнэ.)')
    page7_describe = fields.Text(u'7 - Тэмдэглэл. (Бараа материалын өртгийг тодорхойлоход ашигласан арга, бараа '
                                 u'материалын бүртгэлийн систем, өртөг болон цэвэр боломжит үнийн аль багыг сонгох '
                                 u'аргын талаар тайлбар, тэмдэглэл хийнэ.)')
    page9_describe = fields.Text(u'9 - Тэмдэглэл (Үндсэн хөрөнгийн анги бүрийн хувьд ашигласан хэмжилтийн суурь, '
                                 u'элэгдэл тооцох арга, ашиглалтын хугацаа, дахин үнэлсэн бол дахин үнэлгээ хүчинтэй '
                                 u'болсон хугацаа, хараах бус үнэлгээчин үнэлсэн эсэх талаар, үндсэн хөрөнгийн дахин '
                                 u'ангилал, түүний, шалтгаан, бусад тайлбар тэмдэглэлийг хийнэ)')
    page10_describe = fields.Text(u'10 - Тэмдэглэл (Биет бус хөрөнгийн анги бүрийн хувьд ашигласан хэмжилтийн суурь, '
                                  u'хорогдол тооцох арга, ашиглалтын хугацаа, дахин үнэлсэн бол дахин үнэлгээ хүчинтэй '
                                  u'болсон хугацаа, хараат бус үнэлгээчин үнэлсэн эсэх, бусад биет бус хөрөнгийн '
                                  u'бүрэлдэхүүн болон бусад тайлбар тэмдэглэлийг хийнэ)')
    page12_describe = fields.Text(u'12 - Тэмдэглэл.(Биологийн хөрөнгийн хэмжилтийн суурь болон бусад тайлбар, '
                                  u'тэмдэглэлийг хийнэ.) ')
    page13_describe = fields.Text(u'13 - Тэмдэглэл.(Урт хугацаат хөрөнгө оруулалттай холбоотой бий болсон олз, '
                                  u'гарзын дүн, бүртгэсэн аргыг тодруулна. Охин компани, хамтын хяналттай аж '
                                  u'ахуйн нэгж, хараат компанид оруулсан хөрөнгө оруулалтыг НББОУС-27 Нэгтгэсэн '
                                  u'болон тусдаа санхүүгийн тайлан -ийн дагуу тодруулна.) ')
    page14_describe = fields.Text(u'14 - Тэмдэглэл.(Хөрөнгө оруулалтын зориулалттай үл хөдлөх хөрөнгийн хувьд '
                                  u'ашигласан хэмжилтийн суурь: бодит үнэ цэнийн загвар ашигладаг бол бодит үнэ '
                                  u'цэнийн загвар ашигладаг бол бодит үнэ цэнийг тодорхойлоход ашигласан арга, бодит '
                                  u'үнэ цэнийн тохируулгаас үүссэн олз, гарз; хэрэв түрээслэдэг бол түрээсийн орлого, '
                                  u'түрээсэлсэн хөрөнгөтэй холбоотой гарсан зардлууд; Хэрэв өртгийн загвар ашигладаг '
                                  u'бол хөрөнгийн ашиглалтын хугацаа, элэгдэл тооцох арга болон НББОУС-40 Хөрөнгө '
                                  u'оруулалтын зориулалттай үл хөдлөх хөрөнгө-д заасны дагуу бусад тодруулгыг хийнэ ) ')
    page16_describe = fields.Text(u'16 - Тэмдэглэл. / Урт хугацаат нөөцийн дүнг тодруулна. Нөөцийн төрлөөр '
                                  u'тайлбар , тэмдэглэнэ./')
    page10_ids = fields.One2many('account.disclosure.main.page10', 'main_id', u'10. БИЕТ БУС ХӨРӨНГӨ')
    page11_ids = fields.One2many('account.disclosure.main.page11', 'main_id', u'11. ДУУСААГҮЙ БАРИЛГА')
    page12_ids = fields.One2many('account.disclosure.main.page12', 'main_id', u'12. БИОЛОГИЙН ХӨРӨНГӨ')
    page13_ids = fields.One2many('account.disclosure.main.page13', 'main_id', u'13. УРТ ХУГАЦААТ ХӨРӨНГӨ ОРУУЛАЛТ')
    page15_ids = fields.One2many('account.disclosure.main.page15', 'main_id', u'15. БУСАД ЭРГЭЛТИЙН БУС ХӨРӨНГӨ')
    page16_1_ids = fields.One2many('account.disclosure.main.page16_1', 'main_id', u'16.1 Дансны өглөг')
    page16_2_ids = fields.One2many('account.disclosure.main.page16_2', 'main_id', u'16.2 Татварын өр ')
    page16_3_ids = fields.One2many('account.disclosure.main.page16_3', 'main_id', u'16.3 Богино хугацаат зээл')
    page16_4_ids = fields.One2many('account.disclosure.main.page16_4', 'main_id',
                                   u'16.4 Богино хугацаат нөөц (өр төлбөр)')
    page16_5_ids = fields.One2many('account.disclosure.main.page16_5', 'main_id',
                                   u'16.5 Бусад богино хугацаат өр төлбөр')
    page16_6_ids = fields.One2many('account.disclosure.main.page16_6', 'main_id',
                                   u'16.6 Урт хугацаат зээл болон бусад урт хугацаат өр төлбөр')
    page17_1_ids = fields.One2many('account.disclosure.main.page17_1', 'main_id', u'17.1 Өмч')
    page17_2_ids = fields.One2many('account.disclosure.main.page17_2', 'main_id',
                                   u'17.2 Хөрөнгийн дахин үнэлгээний нэмэгдэл')
    page25_ids = fields.One2many('account.disclosure.main.page25', 'main_id', u'25. ХӨРӨНГӨ ОРУУЛАЛТ')

    @api.model
    def default_get(self, fields_list):
        res = super(AccountDisclosureMain, self).default_get(fields_list)

        page3_list = []
        page4_1_list = []
        page4_2_list = []
        page4_3_list = []
        page17_3_list = []
        page18_1_list = []
        page18_2_list = []
        page18_3_list = []
        page18_4_list = []
        page19_2_list = []
        page19_3_list = []
        page20_1_list = []
        page20_2_list = []
        page20_3_list = []
        page21_list = []
        page22_1_list = []
        page22_2_list = []

        # 3 Гадаад валютын хөрвүүлэлтийн нөөц
        page3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Касс дахь мөнгө'}))
        page3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Банкин дахь мөнгө'}))
        page3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Мөнгө түүнтэй адилтгах хөрөнгө'}))

        # 17.3 Гадаад валютын хөрвүүлэлтийн нөөц
        page17_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Гадаад үйл ажиллагааны хөрвүүлэлтээс үүссэн зөрүү'}))
        page17_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бүртгэлийн валютыг толилуулгын валют руу хөрвүүлснээс үүссэн зөрүү'}))
        page17_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад'}))

        # 18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ
        page18_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Борлуулалтын орлого:'}))
        page18_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бараа, бүтээгдэхүүн борлуулсны орлого'}))
        page18_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ажил, үйлчилгээ борлуулсны орлого'}))
        page18_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Борлуулалтын буцаалт, хөнгөлөлт, үнийн бууралт /-/'}))
        page18_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Борлуулсан бүтээгдэхүүний өртөг:'}))
        page18_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Борлуулсан бараа, борлуулалтын өртөг'}))
        page18_4_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Борлуулсан ажил, үйлчилгээний өртөг'}))

        # 19.2 Гадаад валютын ханшийн зөрүүний олз, гарз
        page19_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Мөнгөн хөрөнгийн үлдэгдэлд хийсэн ханшийн тэгшитгэлийн ханшийн зөрүү'}))
        page19_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эргэлийн авлага, өр төлбөртэй холбоотой үүссэн ханшийн зөрүү'}))
        page19_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эргэлийн бус авлага, өр төлбөртэй холбоотой үүссэн ханшийн зөрүү'}))
        page19_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Валютын арилжаанаас үүссэн олз/гарз'}))

        # 19.3 Бусад ашиг (алдагдал)
        page19_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хөрөнгийн үнэ цэнийн бууралтын гарз'}))
        page19_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'ХОЗҮХХ-ийн бодит үнэ цэнийн өөрчлөлтийн олз, гарз'}))
        page19_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'ХОЗҮХХ данснаас хассаны олз, гарз'}))
        page19_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хөрөнгийн дахин үнэлгээний олз, гарз'}))
        page19_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хөрөнгийн үнэ цэнийн бууралтын гарз /гарзын буцаалт/'}))

        # 20.1 Борлуулалт маркетингийн болон ерөнхий ба удирдлагын зардлууд
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ажиллагчдын цалингийн зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Аж ахуйн нэгжээс төлсөн НДШ-ийн зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Албан татвар, төлбөр, хураамжийн зардал '}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Томилолтын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бичиг хэргийн зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Шуудан холбооны зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Мэргэжлийн үйлчилгээний зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Сургалтын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Сонин сэтгүүл захиалгын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Даатгалын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ашиглалтын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Засварын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Элэгдэл, хорогдлын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Түрээсийн зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Харуул хамгаалалтын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Цэвэрлэгээ үйлчилгээний зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Тээврийн зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Шатахууны зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хүлээн авалтын зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Зар сурталчилгааны зардал'}))
        page20_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад'}))

        # 20.2 Бусад зардал
        page20_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Алданги торгуулийн зардал'}))
        page20_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хандивийн зардал'}))
        page20_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Найдваргүй авлагын зардал'}))
        page20_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад'}))

        # 20.3 Цалингийн зардал
        page20_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үйлдвэрлэл, үйлчлгээний'}))
        page20_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Борлуулалт, маркетингийн'}))
        page20_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ерөнхий ба удирдлагын'}))

        # 21. ОРЛОГЫН ТАТВАРЫН ЗАРДАЛ
        page21_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Тайлангийн үеийн орлогын татварын зардал'}))
        page21_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хойшлогдсон татварын зардал /орлого/'}))

        # 22.1 Толгой компани, хамгийн дээд хяналт тавигч компани, хувь хүний талаарх мэдээлэл****
        page22_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Нэр'}))
        page22_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бүртгэгдсэн /оршин суугаа/ улс'}))
        page22_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эзэмшлийн хувь'}))

        # 22.2 Тэргүүлэх удирдлагын бүрэлдэхүүнд олгосон нөхөн олговрын тухай мэдээлэл
        page22_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Богино хугацааны тэтгэмж'}))
        page22_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Урт хугацааны тэтгэмж'}))
        page22_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ажил эрхлэлтийн дараах тэтгэмж'}))
        page22_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ажлаас халагдсаны тэтгэмж'}))
        page22_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хувьцаанд суурилсан төлбөр'}))

        # 25. ХӨРӨНГӨ ОРУУЛАЛТ
        page6_list = []
        page8_list = []
        page25_list = []
        page9_1_list = []
        page9_2_list = []
        page10_list = []
        page16_1_list = []
        page16_2_list = []
        page16_3_list = []
        page16_4_list = []
        page16_6_list = []
        page17_1_list = []
        page17_2_list = []

        # 4. ДАНСНЫ БОЛОН БУСАД АВЛАГА
        page4_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эхний үлдэгдэл'}))
        page4_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Нэмэгдсэн'}))
#         page4_1_list.append((0, 0, {
#             'main_id': self.id,
#             'name': u'Хасагдсан (-):'}))
        page4_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'    Төлөгдсөн '}))
        page4_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'    Найдваргүй болсон'}))
#         page4_1_list.append((0, 0, {
#             'main_id': self.id,
#             'name': u'Эцсийн үлдэгдэл'}))

        page4_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'ААНОАТ-ын авлага'}))
        page4_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'НӨАТ-ын авлага'}))
        page4_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'НДШ-ийн авлага'}))

        page4_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Холбоотой талаас авлага (эргэлтийн хөрөнгөнд хамаарах дүн)'}))
        page4_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ажиллагчдаас авах авлага '}))
        page4_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Ногдол ашгийн авлага'}))
        page4_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хүүний авлага'}))
        page4_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Богино хугацаат авлагын бичиг '}))
        page4_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад талуудаас авах авлага'}))

        # 6. БАРАА МАТЕРИАЛ
        page6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эхний үлдэгдэл (өртгөөр)'}))
        page6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Нэмэгдсэн дүн'}))
        page6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хасагдсан дүн (-)'}))
#         page6_list.append((0, 0, {
#             'main_id': self.id,
#             'name': u'Эцсийн үлдэгдэл (өртгөөр)'}))
        page6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнийн бууралтын гарз (-)'}))
        page6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнийн бууралтын буцаалт'}))
#         page6_list.append((0, 0, {
#             'main_id': self.id,
#             'name': u'Дансны цэвэр дүн*'}))

#         page6_list.append((0, 0, {
#             'main_id': self.id,
#             'name': u'Эхний үлдэгдэл'}))

#         page6_list.append((0, 0, {
#             'main_id': self.id,
#             'name': u'Эцсийн үлдэгдэл'}))
        # 8
        page8_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Урьдчилж төлсөн зардал'}))
        page8_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Урьдчилж төлсөн түрээс, даатгал'}))
        page8_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бэлтгэн нийлүүлэгчид төлсөн урьдчилгаа төлбөр'}))

        # 9. ҮНДСЭН ХӨРӨНГӨ
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эхний үлдэгдэл'}))
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Өөрөө үйлдвэрлэсэн'}))
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Худалдаж авсан'}))
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэ төлбөргүй авсан'}))
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээний нэмэгдэл'}))
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Худалдсан'}))
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэгүй шилжүүлсэн'}))
        page9_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Акталсан'}))
        # 9.2
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эхний үлдэгдэл'}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үндсэн хөрөнгө дахин ангилсан '}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үндсэн хөрөнгө, ХОЗҮХХ* хооронд дахин ангилсан '}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Байгуулсан элэгдэл '}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээгээр нэмэгдсэн '}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэ цэнийн бууралтын буцаалт'}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Данснаас хассан хөрөнгийн элэгдэл'}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээгээр хасагдсан '}))
        page9_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэ цэнийн бууралт'}))

        # 10. БИЕТ БУС ХӨРӨНГӨ
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эхний үлдэгдэл'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Өөрөө үйлдвэрлэсэн'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Худалдаж авсан'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэ төлбөргүй авсан'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээний нэмэгдэл'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Худалдсан'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэгүй шилжүүлсэн'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Акталж, устгасан'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Байгуулсан хорогдол'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээгээр нэмэгдсэн'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэ цэнийн бууралтын буцаалт'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Данснаас хассан хөрөнгийн хорогдол'}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээгээр хасагдсан '}))
        page10_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үнэ цэнийн бууралт'}))

        # 16.1
        page16_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u' - Төлөгдөх хугацаандаа байгаа'}))
        page16_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u' - Хугацаа хэтэрсэн'}))
        # 16.2
        page16_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'ААНОАТ-ын өр'}))
        page16_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'НӨАТ-ын өр'}))
        page16_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'ХХОАТ-ын өр'}))
        page16_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Онцгой албан татварын өр'}))
        page16_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад татварын өр '}))
        # 16.3
        page16_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u' - Төлөгдөх хугацаандаа байгаа'}))
        page16_3_list.append((0, 0, {
            'main_id': self.id,
            'name': u' - Хугацаа хэтэрсэн'}))
        # 16.4
        page16_4_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Баталгаат засварын'}))
        page16_4_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Нөхөн сэргээлтийн'}))
        # 16.6
        page16_6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Урт хугацаат зээлийн дүн'}))
        page16_6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Гадаадын байгууллагаас шууд авсан зээл'}))
        page16_6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Гадаадын байгууллагаас дамжуулан авсан зээл'}))
        page16_6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дотоодын эх үүсвэрээс авсан зээл'}))
        page16_6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад урт хугацаат өр төлбөрийн дүн'}))
        page16_6_list.append((0, 0, {
            'main_id': self.id,
            'name': u'(Гадаад, дотоодын зах зээлд гаргасан бонд, өрийн бичиг)'}))
        # 17.1

        page17_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эхний үлдэгдэл'}))

        page17_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Нэмэгдсэн '}))

        page17_1_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Хасагдсан (-)'}))

        # 17.2

        page17_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Эхний үлдэгдэл'}))
        page17_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээний нэмэгдлийн зөрүү '}))
        page17_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлсэн хөрөнгийн үнэ цэнийн бууралтын гарзын буцаалт **'}))
        page17_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээний нэмэгдлийн зөрүү '}))
        page17_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлгээний нэмэгдлийн хэрэгжсэн дүн'}))
        page17_2_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Дахин үнэлсэн хөрөнгийн үнэ цэнийн бууралтын гарз***'}))

        # 1 МӨНГӨ БА ТҮҮНТЭЙ АДИЛТГАХ ХӨРӨНГӨ
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Газрын сайжруулалт'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Барилга байгууламж'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үүнээс: Орон сууцны барилга'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Авто зам'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Машин, тоног төхөөрөмж'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Тээврийн хэрэгсэл'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Тавилга эд хогшил'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Компьютер,бусад хөрөнгө'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Биологийн хөрөнгө'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад биет хөрөнгө'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үүнээс:ХОЗҮХХ'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Зохиогчийн эрх'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Компьютер,программ хангамж'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үүнээс:Прогром хангамж'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Мэдээллийн сан'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Патент'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Барааны тэмдэг'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Тусгай зөвшөөрөл'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Газар эзэмших эрх'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Бусад биет хөрөнгө'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үүнээс:Зураг төсвийн ажил,ТЭЗҮ боловсруулах туршилт судалгаа'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Үүнээс: Биет хөрөнгө'}))
        page25_list.append((0, 0, {
            'main_id': self.id,
            'name': u'Биет бус хөрөнгө '}))
#         print ('page3_list========== ',page3_list)
        res.update({
            'page3_ids': page3_list,
            'page8_ids': page8_list,
            'page17_3_ids': page17_3_list,
            'page18_1_ids': page18_1_list,
            'page18_2_ids': page18_2_list,
            'page18_3_ids': page18_3_list,
            'page18_4_ids': page18_4_list,
            'page19_2_ids': page19_2_list,
            'page19_3_ids': page19_3_list,
            'page20_1_ids': page20_1_list,
            'page20_2_ids': page20_2_list,
            'page20_3_ids': page20_3_list,
            'page21_ids': page21_list,
            'page22_1_ids': page22_1_list,
            'page22_2_ids': page22_2_list,
            'page25_ids': page25_list,
            'page4_1_ids': page4_1_list,
            'page4_2_ids': page4_2_list,
            'page4_3_ids': page4_3_list,
            'page6_ids': page6_list,
            'page9_1_ids': page9_1_list,
            'page9_2_ids': page9_2_list,
            'page10_ids': page10_list,
            'page16_1_ids': page16_1_list,
            'page16_2_ids': page16_2_list,
            'page16_3_ids': page16_3_list,
            'page16_4_ids': page16_4_list,
            'page16_6_ids': page16_6_list,
            'page17_1_ids': page17_1_list,
            'page17_2_ids': page17_2_list,
        })
        return res

    
    def compute_with_account(self, account_id, type, date_type, move_type):
        if account_id:
            for this in self:
                balance = 0.0
                if date_type == 'first':
                    if account_id:
                        for account in account_id:
                            initial_bal = self.env['account.move.line'].get_initial_balance(
                                self.company_id.id,
                                [account.id],
                                this.fiscalyear_id.date_start,
                                move_type)
                            if initial_bal:
                                if 'debit' == type:
                                    balance += initial_bal[0]['start_debit']
                                elif 'credit' == type:
                                    balance += initial_bal[0]['start_credit']
                                elif 'all' == type:
                                    balance += initial_bal[0]['start_debit'] - initial_bal[0]['start_credit']
                        return balance
                elif date_type == 'in':
                    if account_id:
                        for account in account_id:
                            initial_bal = self.env['account.move.line'].get_balance(self.company_id.id, [account.id],
                                                                                    this.fiscalyear_id.date_start,
                                                                                    this.fiscalyear_id.date_stop,
                                                                                    move_type)
                            
                            if initial_bal:
                                if type == 'debit':
                                    balance += initial_bal[0]['debit']
                                elif type == 'credit':
                                    balance += initial_bal[0]['credit']
                                elif type == 'all':
                                    balance += initial_bal[0]['credit'] - initial_bal[0]['debit']
                        return balance
                elif date_type == 'last':
                    if account_id:
                        for account in account_id:
                            initial_bal = self.env['account.move.line'].get_initial_balance(
                                self.company_id.id,
                                [account.id],
                                this.fiscalyear_id.date_stop,
                                move_type)
                            if initial_bal:
                                if type == 'debit':
                                    balance += initial_bal[0]['start_debit']
                                elif type == 'credit':
                                    balance += initial_bal[0]['start_credit']
                                elif type == 'all':
                                    balance += initial_bal[0]['start_debit'] - initial_bal[0]['start_credit']
                        return balance

    
    def compute(self):
        move_type = False
        if self.type_n == '1':
            move_type = 'posted'
        company = self.company_id
        report_option_id = self.env['account.disclosure.report.main'].search([('company_id', '=', company.id)])

        report_table1_obj = report_option_id.report_table1_ids
        report_table2_obj = report_option_id.report_table2_ids
        report_table3_obj = report_option_id.report_table3_ids
        report_table4_obj = report_option_id.report_table4_ids
        report_table5_obj = report_option_id.report_table5_ids
        report_table6_1_obj = report_option_id.report_table6_1_ids
        report_table6_2_obj = report_option_id.report_table6_2_ids
        report_table7_1_obj = report_option_id.report_table7_1_ids
        report_table7_2_obj = report_option_id.report_table7_2_ids

        report_table8_2_obj = report_option_id.report_table8_2_ids
        # 3
        val_3_1 = self.page3_ids.search([('name', 'ilike', 'Касс дахь мөнгө'), ('main_id', '=', self.id)])

        page3_1_first_income_amount = self.compute_with_account(report_table1_obj.search(
            [('table1_1', 'ilike', 'Касс дахь мөнгө'), ('main_id', '=', report_option_id.id)]).table1_2, 'all', 'first',
                                                                move_type)
#         print ('page3_1_first_income_amount ',page3_1_first_income_amount)
        page3_1_last_expense_amount = self.compute_with_account(report_table1_obj.search(
            [('table1_1', 'ilike', 'Касс дахь мөнгө'), ('main_id', '=', report_option_id.id)]).table1_2, 'all', 'last',
                                                                move_type)
#         print ('page3_1_last_expense_amount' ,page3_1_last_expense_amount)
#         print ('val_3_1 ',val_3_1)
        val_3_1.write({'income': page3_1_first_income_amount,
                       'expense': page3_1_last_expense_amount})

        val_3_2 = self.page3_ids.search([('name', 'ilike', 'Банкин дахь мөнгө'), ('main_id', '=', self.id)])

        page3_2_first_income_amount = self.compute_with_account(report_table1_obj.search(
            [('table1_1', 'ilike', 'Банкин дахь мөнгө'), ('main_id', '=', report_option_id.id)]).table1_2, 'all',
                                                                'first', move_type)
        page3_2_last_expense_amount = self.compute_with_account(report_table1_obj.search(
            [('table1_1', 'ilike', 'Банкин дахь мөнгө'), ('main_id', '=', report_option_id.id)]).table1_2, 'all',
                                                                'last', move_type)
        val_3_2.write({'income': page3_2_first_income_amount,
                       'expense': page3_2_last_expense_amount})

        val_3_3 = self.page3_ids.search(
            [('name', 'ilike', 'Мөнгө түүнтэй адилтгах хөрөнгө'), ('main_id', '=', self.id)])

        page3_3_first_income_amount = self.compute_with_account(report_table1_obj.search(
            [('table1_1', 'ilike', 'Мөнгө түүнтэй адилтгах хөрөнгө'), ('main_id', '=', report_option_id.id)]).table1_2,
                                                                'all', 'first', move_type)
        page3_3_last_expense_amount = self.compute_with_account(report_table1_obj.search(
            [('table1_1', 'ilike', 'Мөнгө түүнтэй адилтгах хөрөнгө'), ('main_id', '=', report_option_id.id)]).table1_2,
                                                                'all', 'last', move_type)
        val_3_3.write({'income': page3_3_first_income_amount,
                       'expense': page3_3_last_expense_amount})
        # 3.1
        #         Эхний үлдэгдэл
        first = self.page4_1_ids.search([('name', 'ilike', 'Эхний үлдэгдэл'), ('main_id', '=', self.id)])
        add = self.page4_1_ids.search([('name', 'ilike', 'Нэмэгдсэн'), ('main_id', '=', self.id)])
        paid = self.page4_1_ids.search([('name', 'ilike', 'Төлөгдсөн'), ('main_id', '=', self.id)])
        failed = self.page4_1_ids.search([('name', 'ilike', 'Найдваргүй болсон'), ('main_id', '=', self.id)])
        page4_1_first_income_amount = self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Дансны авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'first',
                                                                move_type) or 0
        page4_1_first_expense_amount = self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Найдваргүй авлагын хасагдуулга'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                                 'all', 'first', move_type) or 0
        first.write({'income': page4_1_first_income_amount or 0,
                     'expense': page4_1_first_expense_amount or 0})
        #         Нэмэгдсэн
        page4_1_add_debit_income_amount = self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Дансны авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'debit', 'in',
                                                                    move_type) or 0
        page4_1_add_debit_expense_amount = self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Найдваргүй авлагын хасагдуулга'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                                     'debit', 'in', move_type) or 0
        amount = page4_1_add_debit_income_amount - page4_1_add_debit_expense_amount if page4_1_add_debit_income_amount and page4_1_add_debit_expense_amount else 0
        add.write({'income': page4_1_add_debit_income_amount,
                   'expense': page4_1_add_debit_expense_amount,
#                    'amount': amount or 0
                   })
        total_income1 = 0
        total_expense1 = 0
        total_income2 = 0
        total_expense2 = 0

        # Singleton-ий алдаа гарч байсан тул давталтад оруулав
        for obj in paid:
            if obj.income:
                total_income1 = obj.income
            elif obj.expense:
                total_expense1 = obj.expense

        if failed.income:
            total_income2 = failed.income
        elif failed.expense:
            total_expense2 = failed.expense

        paid.write({'income': total_income1 or 0,
                    'expense': total_expense1 or 0,
#                     'amount': total_income1 - total_expense1 if total_income1 and total_expense1 else 0
                    })

        failed.write({'income': total_income2 or 0,
                      'expense': total_expense2 or 0,
#                       'amount': total_income2 - total_expense2 if total_income2 and total_expense2 else 0
                      })
        # 4.2
        annoat = self.page4_2_ids.search([('name', 'ilike', 'ААНОАТ-ын авлага'), ('main_id', '=', self.id)])
        noat = self.page4_2_ids.search([('name', 'ilike', 'НӨАТ-ын авлага'), ('main_id', '=', self.id)])
        ndsh = self.page4_2_ids.search([('name', 'ilike', 'НДШ-ийн авлага'), ('main_id', '=', self.id)])

        annoat.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'ААНОАТ-ын авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'all',
                                                          'first', move_type),
                      'expense': self.compute_with_account(report_table2_obj.search(
                          [('table2_1', 'ilike', 'ААНОАТ-ын авлага'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                           'all', 'last', move_type)})
        noat.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'НӨАТ-ын авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'first',
                                                        move_type),
                    'expense': self.compute_with_account(report_table2_obj.search(
                        [('table2_1', 'ilike', 'НӨАТ-ын авлага'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                         'all', 'last', move_type)})
        ndsh.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'НДШ-ийн авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'first',
                                                        move_type),
                    'expense': self.compute_with_account(report_table2_obj.search(
                        [('table2_1', 'ilike', 'НДШ-ийн авлага'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                         'all', 'last', move_type)})
        # 4.3
        table4_3_6 = self.page4_3_ids.search([('name', 'ilike', 'Холбоотой талаас авлага'), ('main_id', '=', self.id)])
        table4_3_7 = self.page4_3_ids.search([('name', 'ilike', 'Ажиллагчдаас авах авлага'), ('main_id', '=', self.id)])
        table4_3_8 = self.page4_3_ids.search([('name', 'ilike', 'Ногдол ашгийн авлага'), ('main_id', '=', self.id)])
        table4_3_9 = self.page4_3_ids.search([('name', 'ilike', 'Хүүний авлага'), ('main_id', '=', self.id)])
        table4_3_10 = self.page4_3_ids.search(
            [('name', 'ilike', 'Богино хугацаат авлагын бичиг'), ('main_id', '=', self.id)])
        table4_3_11 = self.page4_3_ids.search(
            [('name', 'ilike', 'Бусад талуудаас авах авлага'), ('main_id', '=', self.id)])

        table4_3_6.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Холбоотой талаас авлага (эргэлтийн хөрөнгөнд хамаарах дүн)'),
             ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'first', move_type),
                          'expense': self.compute_with_account(report_table2_obj.search(
                              [('table2_1', 'ilike', 'Холбоотой талаас авлага (эргэлтийн хөрөнгөнд хамаарах дүн)'),
                               ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'last', move_type)})

        table4_3_7.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Ажиллагчдаас авах авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'all',
                                                              'first', move_type),
                          'expense': self.compute_with_account(report_table2_obj.search(
                              [('table2_1', 'ilike', 'Ажиллагчдаас авах авлага'),
                               ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'last', move_type)})

        table4_3_8.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Ногдол ашгийн авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'all',
                                                              'first', move_type),
                          'expense': self.compute_with_account(report_table2_obj.search(
                              [('table2_1', 'ilike', 'Ногдол ашгийн авлага'),
                               ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'last', move_type)})

        table4_3_9.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Хүүний авлага'), ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'first',
                                                              move_type),
                          'expense': self.compute_with_account(report_table2_obj.search(
                              [('table2_1', 'ilike', 'Хүүний авлага'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                               'all', 'last', move_type)})

        table4_3_10.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Богино хугацаат авлагын бичиг'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                               'all', 'first', move_type),
                           'expense': self.compute_with_account(report_table2_obj.search(
                               [('table2_1', 'ilike', 'Богино хугацаат авлагын бичиг'),
                                ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'last', move_type)})

        table4_3_11.write({'income': self.compute_with_account(report_table2_obj.search(
            [('table2_1', 'ilike', 'Бусад талуудаас авах авлага'), ('main_id', '=', report_option_id.id)]).table2_2,
                                                               'all', 'first', move_type),
                           'expense': self.compute_with_account(report_table2_obj.search(
                               [('table2_1', 'ilike', 'Бусад талуудаас авах авлага'),
                                ('main_id', '=', report_option_id.id)]).table2_2, 'all', 'last', move_type)})

        for report_object in report_table3_obj:
            self.page5_ids.create({
                'name': report_object.table3_1,
                'main_id': self.id,
            })

        table5_ids_list = []
        for obj in self.page8_ids:
            if obj.table5_id:
                table5_ids_list.append(obj.name)

        for report_obj in report_table5_obj:
            if report_obj.table5_1 not in table5_ids_list:
                self.page8_ids.create({
                    'name': report_obj.table5_1,
                    'main_id': self.id,
                    'table5_id':report_obj.id,
                    })


        table6_1 = self.page6_ids.search([('name', 'ilike', 'Эхний үлдэгдэл'), ('main_id', '=', self.id)])
        table6_2 = self.page6_ids.search([('name', 'ilike', 'Нэмэгдсэн дүн'), ('main_id', '=', self.id)])
        table6_3 = self.page6_ids.search([('name', 'ilike', 'Хасагдсан дүн'), ('main_id', '=', self.id)])
        table6_1.write({'raw_materials': self.compute_with_account(report_table4_obj.search(
            [('table4_1', 'ilike', 'Түүхий эд материал'), ('main_id', '=', report_option_id.id)]).table4_2, 'all',
                                                                   'first', move_type),
                        'mrp': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Дуусаагүй үйлдвэрлэл'),
                             ('main_id', '=', report_option_id.id)]).table4_2, 'all', 'first', move_type),
                        'production': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Бэлэн бүтээгдэхүүн'),
                             ('main_id', '=', report_option_id.id)]).table4_2, 'all', 'first', move_type),
                        'product': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Бараа'), ('main_id', '=', report_option_id.id)]).table4_2, 'all',
                                                             'first', move_type),
                        'supply_materials': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Хангамж'), ('main_id', '=', report_option_id.id)]).table4_2, 'all',
                                                                      'first', move_type),
                        })
# Нэмэгдсэн дүн
        table6_2.write({'raw_materials': self.compute_with_account(report_table4_obj.search(
            [('table4_1', 'ilike', 'Түүхий эд материал'), ('main_id', '=', report_option_id.id)]).table4_2, 'debit',
                                                                   'in', move_type),
                        'mrp': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Дуусаагүй үйлдвэрлэл'),
                             ('main_id', '=', report_option_id.id)]).table4_2, 'debit', 'in', move_type),
                        'production': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Бэлэн бүтээгдэхүүн'),
                             ('main_id', '=', report_option_id.id)]).table4_2, 'debit', 'in', move_type),
                        'product': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Бараа'), ('main_id', '=', report_option_id.id)]).table4_2, 'debit',
                                                             'in', move_type),
                        'supply_materials': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Хангамж'), ('main_id', '=', report_option_id.id)]).table4_2,
                                                                      'debit', 'in', move_type),
                        })
# Хасагдсан дүн
        table6_3.write({'raw_materials': self.compute_with_account(report_table4_obj.search(
            [('table4_1', 'ilike', 'Түүхий эд материал'), ('main_id', '=', report_option_id.id)]).table4_2, 'credit',
                                                                   'in', move_type),
                        'mrp': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Дуусаагүй үйлдвэрлэл'),
                             ('main_id', '=', report_option_id.id)]).table4_2, 'credit', 'in', move_type),
                        'production': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Бэлэн бүтээгдэхүүн'),
                             ('main_id', '=', report_option_id.id)]).table4_2, 'credit', 'in', move_type),
                        'product': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Бараа'), ('main_id', '=', report_option_id.id)]).table4_2, 'credit',
                                                             'in', move_type),
                        'supply_materials': self.compute_with_account(report_table4_obj.search(
                            [('table4_1', 'ilike', 'Хангамж'), ('main_id', '=', report_option_id.id)]).table4_2,
                                                                      'credit', 'in', move_type),
                        })

        table8_1 = self.page8_ids.search([('name', 'ilike', 'Урьдчилж төлсөн зардал'), ('main_id', '=', self.id)])
        table8_2 = self.page8_ids.search(
            [('name', 'ilike', 'Урьдчилж төлсөн түрээс, даатгал'), ('main_id', '=', self.id)])
        table8_3 = self.page8_ids.search(
            [('name', 'ilike', 'Бэлтгэн нийлүүлэгчид төлсөн урьдчилгаа төлбөр'), ('main_id', '=', self.id)])

        table8_1.write({'first_amount': self.compute_with_account(report_table5_obj.search(
            [('table5_1', 'ilike', 'Урьдчилж төлсөн зардал'), ('main_id', '=', report_option_id.id)]).table5_2, 'all',
                                                                  'first', move_type),
                        'last_amount': self.compute_with_account(report_table5_obj.search(
                            [('table5_1', 'ilike', 'Урьдчилж төлсөн зардал'),
                             ('main_id', '=', report_option_id.id)]).table5_2, 'all', 'last', move_type)})
        table8_2.write({'first_amount': self.compute_with_account(report_table5_obj.search(
            [('table5_1', 'ilike', 'Урьдчилж төлсөн түрээс, даатгал'), ('main_id', '=', report_option_id.id)]).table5_2,
                                                                  'all', 'first', move_type),
                        'last_amount': self.compute_with_account(report_table5_obj.search(
                            [('table5_1', 'ilike', 'Урьдчилж төлсөн түрээс, даатгал'),
                             ('main_id', '=', report_option_id.id)]).table5_2, 'all', 'last', move_type)})
        table8_3.write({'first_amount': self.compute_with_account(report_table5_obj.search(
            [('table5_1', 'ilike', 'Бэлтгэн нийлүүлэгчид төлсөн урьдчилгаа төлбөр'),
             ('main_id', '=', report_option_id.id)]).table5_2, 'all', 'first', move_type),
                        'last_amount': self.compute_with_account(report_table5_obj.search(
                            [('table5_1', 'ilike', 'Бэлтгэн нийлүүлэгчид төлсөн урьдчилгаа төлбөр'),
                             ('main_id', '=', report_option_id.id)]).table5_2, 'all', 'last', move_type)})
        # 9
        table9_1 = self.page9_1_ids.search([('name', 'ilike', 'Эхний үлдэгдэл'), ('main_id', '=', self.id)])
        table9_1.write({
            'gazriin_saijruulalt': self.compute_with_account(report_table6_1_obj.search(
            [('table6_1_1', 'ilike', 'Газрын сайжруулалт'), ('main_id', '=', report_option_id.id)]).table6_1_2, 'all',
                                                                   'first', move_type),
            'barilga': self.compute_with_account(report_table6_1_obj.search(
            [('table6_1_1', 'ilike', 'Барилга, байгууламж'), ('main_id', '=', report_option_id.id)]).table6_1_2, 'all',
                                                                   'first', move_type),
            'mashin_totu': self.compute_with_account(report_table6_1_obj.search(
            [('table6_1_1', 'ilike', 'Машин, тоног төхөөрөмж'), ('main_id', '=', report_option_id.id)]).table6_1_2, 'all',
                                                                   'first', move_type),
            'teevriin': self.compute_with_account(report_table6_1_obj.search(
            [('table6_1_1', 'ilike', 'Тээврийн хэрэгсэл'), ('main_id', '=', report_option_id.id)]).table6_1_2, 'all',
                                                                   'first', move_type),
            'tavilga': self.compute_with_account(report_table6_1_obj.search(
            [('table6_1_1', 'ilike', 'Тавилга эд хогшил'), ('main_id', '=', report_option_id.id)]).table6_1_2, 'all',
                                                                   'first', move_type),
            'computer': self.compute_with_account(report_table6_1_obj.search(
            [('table6_1_1', 'ilike', 'Компьютер, бусад хэрэгсэл'), ('main_id', '=', report_option_id.id)]).table6_1_2, 'all',
                                                                   'first', move_type),
            'busad': self.compute_with_account(report_table6_1_obj.search(
            [('table6_1_1', 'ilike', 'Бусад үндсэн хөрөнгө'), ('main_id', '=', report_option_id.id)]).table6_1_2, 'all',
                                                                   'first', move_type), })
        table9_2 = self.page9_2_ids.search([('name', 'ilike', 'Эхний үлдэгдэл'), ('main_id', '=', self.id)])
        table9_2.write({
            'gazriin_saijruulalt': self.compute_with_account(report_table6_2_obj.search(
            [('table6_2_1', 'ilike', 'Газрын сайжруулалт'), ('main_id', '=', report_option_id.id)]).table6_2_2, 'all',
                                                                   'first', move_type),
            'barilga': self.compute_with_account(report_table6_2_obj.search(
            [('table6_2_1', 'ilike', 'Барилга, байгууламж'), ('main_id', '=', report_option_id.id)]).table6_2_2, 'all',
                                                                   'first', move_type),
            'mashin_totu': self.compute_with_account(report_table6_2_obj.search(
            [('table6_2_1', 'ilike', 'Машин, тоног төхөөрөмж'), ('main_id', '=', report_option_id.id)]).table6_2_2, 'all',
                                                                   'first', move_type),
            'teevriin': self.compute_with_account(report_table6_2_obj.search(
            [('table6_2_1', 'ilike', 'Тээврийн хэрэгсэл'), ('main_id', '=', report_option_id.id)]).table6_2_2, 'all',
                                                                   'first', move_type),
            'tavilga': self.compute_with_account(report_table6_2_obj.search(
            [('table6_2_1', 'ilike', 'Тавилга эд хогшил'), ('main_id', '=', report_option_id.id)]).table6_2_2, 'all',
                                                                   'first', move_type),
            'computer': self.compute_with_account(report_table6_2_obj.search(
            [('table6_2_1', 'ilike', 'Компьютер, бусад хэрэгсэл'), ('main_id', '=', report_option_id.id)]).table6_2_2, 'all',
                                                                   'first', move_type),
            'busad': self.compute_with_account(report_table6_2_obj.search(
            [('table6_2_1', 'ilike', 'Бусад үндсэн хөрөнгө'), ('main_id', '=', report_option_id.id)]).table6_2_2, 'all',
                                                                   'first', move_type), })
        # 10
        table10_1 = self.page10_1_ids.search([('name', 'ilike', 'Эхний үлдэгдэл'), ('main_id', '=', self.id)])
        table10_1.write({
            'zohiogchiin_erh': self.compute_with_account(report_table7_1_obj.search(
            [('table7_1_1', 'ilike', 'Зохиогчийн эрх'), ('main_id', '=', report_option_id.id)]).table7_1_2, 'all',
                                                                   'first', move_type),
            'computer': self.compute_with_account(report_table7_1_obj.search(
            [('table7_1_1', 'ilike', 'Компьютерийн П'), ('main_id', '=', report_option_id.id)]).table7_1_2, 'all',
                                                                   'first', move_type),
            'patent': self.compute_with_account(report_table7_1_obj.search(
            [('table7_1_1', 'ilike', 'Патент'), ('main_id', '=', report_option_id.id)]).table7_1_2, 'all',
                                                                   'first', move_type),
            'baraanii_temdeg': self.compute_with_account(report_table7_1_obj.search(
            [('table7_1_1', 'ilike', 'Барааны тэмдэг'), ('main_id', '=', report_option_id.id)]).table7_1_2, 'all',
                                                                   'first', move_type),
            'tusgai_zovshoorol': self.compute_with_account(report_table7_1_obj.search(
            [('table7_1_1', 'ilike', 'Тусгай зөвшөөрөл'), ('main_id', '=', report_option_id.id)]).table7_1_2, 'all',
                                                                   'first', move_type),
            'gazar': self.compute_with_account(report_table7_1_obj.search(
            [('table7_1_1', 'ilike', 'Газар эзэмших эрх'), ('main_id', '=', report_option_id.id)]).table7_1_2, 'all',
                                                                   'first', move_type),
            'biet_bus_horongo': self.compute_with_account(report_table7_1_obj.search(
            [('table7_1_1', 'ilike', 'Бусад биет бус хөрөнгө'), ('main_id', '=', report_option_id.id)]).table7_1_2, 'all',
                                                                   'first', move_type),
            })
        
        table16_2_1 = self.page16_2_ids.search([('name', 'ilike', 'ААНОАТ-ын өр'), ('main_id', '=', self.id)])
        table16_2_2 = self.page16_2_ids.search([('name', 'ilike', 'НӨАТ-ын өр'), ('main_id', '=', self.id)])
        table16_2_3 = self.page16_2_ids.search([('name', 'ilike', 'ХХОАТ-ын өр'), ('main_id', '=', self.id)])
        table16_2_4 = self.page16_2_ids.search(
            [('name', 'ilike', 'Онцгой албан татварын өр'), ('main_id', '=', self.id)])
        table16_2_5 = self.page16_2_ids.search([('name', 'ilike', 'Бусад татварын өр'), ('main_id', '=', self.id)])

        table16_2_1.write({'first_amount': self.compute_with_account(report_table8_2_obj.search(
            [('table8_2_1', 'ilike', 'ААНОАТ-ын өр'), ('main_id', '=', report_option_id.id)]).table8_2_2, 'all',
                                                                     'first', move_type),
                           'last_amount': self.compute_with_account(report_table8_2_obj.search(
                               [('table8_2_1', 'ilike', 'ААНОАТ-ын өр'),
                                ('main_id', '=', report_option_id.id)]).table8_2_2, 'all', 'last', move_type)})

        table16_2_2.write({'first_amount': self.compute_with_account(report_table8_2_obj.search(
            [('table8_2_1', 'ilike', 'НӨАТ-ын өр'), ('main_id', '=', report_option_id.id)]).table8_2_2, 'all', 'first',
                                                                     move_type),
                           'last_amount': self.compute_with_account(report_table8_2_obj.search(
                               [('table8_2_1', 'ilike', 'НӨАТ-ын өр'),
                                ('main_id', '=', report_option_id.id)]).table8_2_2, 'all', 'last', move_type)})

        table16_2_3.write({'first_amount': self.compute_with_account(report_table8_2_obj.search(
            [('table8_2_1', 'ilike', 'ХХОАТ-ын өр'), ('main_id', '=', report_option_id.id)]).table8_2_2, 'all', 'first',
                                                                     move_type),
                           'last_amount': self.compute_with_account(report_table8_2_obj.search(
                               [('table8_2_1', 'ilike', 'ХХОАТ-ын өр'),
                                ('main_id', '=', report_option_id.id)]).table8_2_2, 'all', 'last', move_type)})

        table16_2_4.write({'first_amount': self.compute_with_account(report_table8_2_obj.search(
            [('table8_2_1', 'ilike', 'Онцгой албан татварын өр'), ('main_id', '=', report_option_id.id)]).table8_2_2,
                                                                     'all', 'first', move_type),
                           'last_amount': self.compute_with_account(report_table8_2_obj.search(
                               [('table8_2_1', 'ilike', 'Онцгой албан татварын өр'),
                                ('main_id', '=', report_option_id.id)]).table8_2_2, 'all', 'last', move_type)})

        table16_2_5.write({'first_amount': self.compute_with_account(report_table8_2_obj.search(
            [('table8_2_1', 'ilike', 'Бусад татварын өр'), ('main_id', '=', report_option_id.id)]).table8_2_2, 'all',
                                                                     'first', move_type),
                           'last_amount': self.compute_with_account(report_table8_2_obj.search(
                               [('table8_2_1', 'ilike', 'Бусад татварын өр'),
                                ('main_id', '=', report_option_id.id)]).table8_2_2, 'all', 'last', move_type)})


#DARMAA ZARDAL
        report_table11_1_obj = report_option_id.report_table11_1_ids
#         print ('report_table11_1_obj ',report_table11_1_obj)
        table20_1_1 = self.page20_1_ids.search([('name', 'ilike', 'Ажиллагчдын цалингийн зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_1.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Ажиллагчдын цалингийн зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', ' '),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})   
        table20_1_2 = self.page20_1_ids.search([('name', 'ilike', 'Аж ахуйн нэгжээс төлсөн НДШ-ийн зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_2.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Аж ахуйн нэгжээс төлсөн НДШ-ийн зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Аж ахуйн нэгжээс төлсөн НДШ-ийн зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})   
        table20_1_3 = self.page20_1_ids.search([('name', 'ilike', 'Албан татвар, төлбөр, хураамжийн зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_3.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Албан татвар, төлбөр, хураамжийн зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', ' '),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})   
        table20_1_4 = self.page20_1_ids.search([('name', 'ilike', 'Томилолтын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_4.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Томилолтын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Томилолтын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})  
        
        table20_1_5 = self.page20_1_ids.search([('name', 'ilike', 'Бичиг хэргийн зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_5.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Бичиг хэргийн зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Бичиг хэргийн зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})                                                

        table20_1_6 = self.page20_1_ids.search([('name', 'ilike', 'Томилолтын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_6.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Шуудан холбооны зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Шуудан холбооны зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_7 = self.page20_1_ids.search([('name', 'ilike', 'Мэргэжлийн үйлчилгээний зардал'), ('main_id', '=', self.id)])
#         print ('able20_1 ',table20_1_1)
        table20_1_7.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Мэргэжлийн үйлчилгээний зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Мэргэжлийн үйлчилгээний зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_8 = self.page20_1_ids.search([('name', 'ilike', 'Сургалтын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_8.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Сургалтын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Сургалтын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_9 = self.page20_1_ids.search([('name', 'ilike', 'Сонин сэтгүүл захиалгын зардал'), ('main_id', '=', self.id)])
#         print ('9able20_1 ',table20_1_1)
        table20_1_9.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Сонин сэтгүүл захиалгын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Сонин сэтгүүл захиалгын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_10 = self.page20_1_ids.search([('name', 'ilike', 'Даатгалын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_10.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Даатгалын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Даатгалын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_11 = self.page20_1_ids.search([('name', 'ilike', 'Ашиглалтын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_11.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Ашиглалтын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Ашиглалтын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_12 = self.page20_1_ids.search([('name', 'ilike', 'Засварын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_12.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Засварын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Засварын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_13 = self.page20_1_ids.search([('name', 'ilike', 'Элэгдэл, хорогдлын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_13.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Элэгдэл, хорогдлын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Томилолтын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_14 = self.page20_1_ids.search([('name', 'ilike', 'Түрээсийн зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_14.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Түрээсийн зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Түрээсийн зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})          
        
        table20_1_15 = self.page20_1_ids.search([('name', 'ilike', 'Тээврийн зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_15.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Тээврийн зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Тээврийн зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})      
        

        table20_1_16 = self.page20_1_ids.search([('name', 'ilike', 'Цэвэрлэгээ үйлчилгээний зардал'), ('main_id', '=', self.id)])
#         print ('tble20_1 ',table20_1_1)
        table20_1_16.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Цэвэрлэгээ үйлчилгээний зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Цэвэрлэгээ үйлчилгээний зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})      
        
        table20_1_17 = self.page20_1_ids.search([('name', 'ilike', 'Шатахууны зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_17.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Шатахууны зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Шатахууны зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})      
        
        table20_1_18 = self.page20_1_ids.search([('name', 'ilike', 'Хүлээн авалтын зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_18.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Хүлээн авалтын зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Хүлээн авалтын зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})      
        
        table20_1_19 = self.page20_1_ids.search([('name', 'ilike', 'Зар сурталчилгааны зардал'), ('main_id', '=', self.id)])
#         print ('table20_1 ',table20_1_1)
        table20_1_19.write({'this_year_sale_amount': self.compute_with_account(report_table11_1_obj.search(
            [('table11_1_1', 'ilike', 'Зар сурталчилгааны зардал'), ('main_id', '=', report_option_id.id)]).table11_1_2, 'debit',
                                                                     'in', move_type),
                           'this_year_management_amount': self.compute_with_account(report_table11_1_obj.search(
                               [('table11_1_1', 'ilike', 'Зар сурталчилгааны зардал'),
                                ('main_id', '=', report_option_id.id)]).table11_1_2, 'all', 'last', move_type)})                  
    
    def action_export(self):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        company = self.company_id

        template_file = get_module_resource('mw_account_disclosure_statement', 'static/src/report',
                                            'disclosure_in_template.xlsx')
        generate_file = get_module_resource('mw_account_disclosure_statement', 'static/src/report',
                                            'disclosure_out_template.xlsx')
        wb = openpyxl.load_workbook(template_file)

        list22_1 = [None] * 3
        i = 0
        for page in self.page22_1_ids:
            list22_1[i] = {}
            list22_1[i]['tc'] = page.head_company
            list22_1[i]['hdhttc'] = page.hamgiin_deed_hyanalt_tavigch_tolgoi_company
            list22_1[i]['hdhthh'] = page.hamgiin_deed_hyanalt_tavigch_huvi_hun
            list22_1[i]['t'] = page.description
            i += 1

        list22_2 = [None] * 5
        i = 0
        for page in self.page22_2_ids:
            list22_2[i] = {}
            list22_2[i]['ib'] = page.initial_balance
            list22_2[i]['eb'] = page.end_balance
            i += 1

        investment = [None] * 23
        i = 0
        for page in self.page25_ids:
            investment[i] = {}
            investment[i]['ib'] = page.initial_balance
            investment[i]['aanoh'] = page.aj_ahui_negjiin_ooriin_horongoor
            investment[i]['uth'] = page.ulsiin_tosviin_horongoor
            investment[i]['onth'] = page.oron_nutgiin_tosviin_horongoor
            investment[i]['bz'] = page.banknii_zeel
            investment[i]['gsho'] = page.gadaadiin_shuud_horongo_oruulalt
            investment[i]['gz'] = page.gadaadiin_zeel
            investment[i]['gbt'] = page.gadaadiin_butsaltgui_tuslamj
            investment[i]['thh'] = page.tosol_hotolbor_handiv
            investment[i]['beu'] = page.busad_eh_uusver
            i += 1

        list22 = [None] * 5
        i = 0
        for page in self.page22_2_ids:
            list22[i] = {}
            list22[i]['ib'] = page.initial_balance
            list22[i]['eb'] = page.end_balance
            i += 1

        investment = [None] * 23
        i = 0
        for page in self.page25_ids:
            investment[i] = {}
            investment[i]['ib'] = page.initial_balance
            investment[i]['aanoh'] = page.aj_ahui_negjiin_ooriin_horongoor
            investment[i]['uth'] = page.ulsiin_tosviin_horongoor
            investment[i]['onth'] = page.oron_nutgiin_tosviin_horongoor
            investment[i]['bz'] = page.banknii_zeel
            investment[i]['gsho'] = page.gadaadiin_shuud_horongo_oruulalt
            investment[i]['gz'] = page.gadaadiin_zeel
            investment[i]['gbt'] = page.gadaadiin_butsaltgui_tuslamj
            investment[i]['thh'] = page.tosol_hotolbor_handiv
            investment[i]['beu'] = page.busad_eh_uusver
            i += 1
        dicts = {
            'company_name': u'Аж ахуйн нэгжийн нэр:  %s' % company.name,
            'date_y_m_d': u'Огноо: %s-12-31' % self.fiscalyear_id.code,
            '{1}': self.intro1,
            '{2}': self.intro2,
            '{page3_1_income}': float(self.page3_ids.search(
                [('name', 'ilike', 'Касс дахь мөнгө'), ('main_id', '=', self.id)]).income) or 0,
            '{page3_1_expense}': self.page3_ids.search(
                [('name', 'ilike', 'Касс дахь мөнгө'), ('main_id', '=', self.id)]).expense,
            
            '{page3_2_income}': self.page3_ids.search(
                [('name', 'ilike', 'Банкин дахь мөнгө'), ('main_id', '=', self.id)]).income,
            '{page3_2_expense}': self.page3_ids.search(
                [('name', 'ilike', 'Банкин дахь мөнгө'), ('main_id', '=', self.id)]).expense,

            '{page3_3_income}': self.page3_ids.search(
                [('name', 'ilike', 'Мөнгө түүнтэй адилтгах хөрөнгө'), ('main_id', '=', self.id)]).income,
            '{page3_3_expense}': self.page3_ids.search(
                [('name', 'ilike', 'Мөнгө түүнтэй адилтгах хөрөнгө'), ('main_id', '=', self.id)]).expense,

            '{page4_1_income}': self.page4_1_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'), ('main_id', '=', self.id)]).income,
            '{page4_1_expense}': self.page4_1_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'), ('main_id', '=', self.id)]).expense,
            '{page4_1_add_income}': self.page4_1_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн'), ('main_id', '=', self.id)]).income,
            '{page4_1_add_expense}': self.page4_1_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн'), ('main_id', '=', self.id)]).expense,
            '{page4_1_paid_income}': self.page4_1_ids.search(
                [('name', 'ilike', 'Төлөгдсөн '), ('main_id', '=', self.id)]).income,
            '{page4_1_paid_expense}': self.page4_1_ids.search(
                [('name', 'ilike', 'Төлөгдсөн '), ('main_id', '=', self.id)]).expense,
            '{page4_1_fail_income}': self.page4_1_ids.search(
                [('name', 'ilike', 'Найдваргүй болсон'), ('main_id', '=', self.id)]).income,
            '{page4_1_fail_expense}': self.page4_1_ids.search(
                [('name', 'ilike', 'Найдваргүй болсон'), ('main_id', '=', self.id)]).expense,

            '{page4_2_1_income}': self.page4_2_ids.search(
                [('name', 'ilike', 'ААНОАТ-ын авлага'), ('main_id', '=', self.id)]).income,
            '{page4_2_1_expense}': self.page4_2_ids.search(
                [('name', 'ilike', 'ААНОАТ-ын авлага'), ('main_id', '=', self.id)]).expense,

            '{page4_2_2_income}': self.page4_2_ids.search(
                [('name', 'ilike', 'НӨАТ-ын авлага'), ('main_id', '=', self.id)]).income,
            '{page4_2_2_expense}': self.page4_2_ids.search(
                [('name', 'ilike', 'НӨАТ-ын авлага'), ('main_id', '=', self.id)]).expense,

            '{page4_2_3_income}': self.page4_2_ids.search(
                [('name', 'ilike', 'НДШ-ийн авлага'), ('main_id', '=', self.id)]).income,
            '{page4_2_3_expense}': self.page4_2_ids.search(
                [('name', 'ilike', 'НДШ-ийн авлага'), ('main_id', '=', self.id)]).expense,

            '{page4_3_1_income}': self.page4_3_ids.search(
                [('name', 'ilike', 'Холбоотой талаас авлага (эргэлтийн хөрөнгөнд хамаарах дүн)'),
                 ('main_id', '=', self.id)]).income,
            '{page4_3_1_expense}': self.page4_3_ids.search(
                [('name', 'ilike', 'Холбоотой талаас авлага (эргэлтийн хөрөнгөнд хамаарах дүн)'),
                 ('main_id', '=', self.id)]).expense,

            '{page4_3_2_income}': self.page4_3_ids.search(
                [('name', 'ilike', 'Ажиллагчдаас авах авлага '), ('main_id', '=', self.id)]).income,
            '{page4_3_2_expense}': self.page4_3_ids.search(
                [('name', 'ilike', 'Ажиллагчдаас авах авлага '), ('main_id', '=', self.id)]).expense,

            '{page4_3_3_income}': self.page4_3_ids.search(
                [('name', 'ilike', 'Ногдол ашгийн авлага'), ('main_id', '=', self.id)]).income,
            '{page4_3_3_expense}': self.page4_3_ids.search(
                [('name', 'ilike', 'Ногдол ашгийн авлага'), ('main_id', '=', self.id)]).expense,

            '{page4_3_4_income}': self.page4_3_ids.search(
                [('name', 'ilike', 'Хүүний авлага'), ('main_id', '=', self.id)]).income,
            '{page4_3_4_expense}': self.page4_3_ids.search(
                [('name', 'ilike', 'Хүүний авлага'), ('main_id', '=', self.id)]).expense,

            '{page4_3_5_income}': self.page4_3_ids.search(
                [('name', 'ilike', 'Богино хугацаат авлагын бичиг'), ('main_id', '=', self.id)]).income,
            '{page4_3_5_expense}': self.page4_3_ids.search(
                [('name', 'ilike', 'Богино хугацаат авлагын бичиг'), ('main_id', '=', self.id)]).expense,

            '{page4_3_6_income}': self.page4_3_ids.search(
                [('name', 'ilike', 'Бусад талуудаас авах авлага'), ('main_id', '=', self.id)]).income,
            '{page4_3_6_expense}': self.page4_3_ids.search(
                [('name', 'ilike', 'Бусад талуудаас авах авлага'), ('main_id', '=', self.id)]).expense,

            '{page4_describe}': self.page4_describe,
            '{page7_describe}': self.page7_describe,
            '{page9_describe}': self.page9_describe,
            '{page10_describe}': self.page10_describe,
            '{page12_describe}': self.page12_describe,
            '{page13_describe}': self.page13_describe,
            '{page14_describe}': self.page14_describe,
            '{page16_describe}': self.page16_describe,

            '{page6_1_1_1}': self.page6_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл (өртгөөр)'), ('main_id', '=', self.id)]).raw_materials,
            '{page6_1_2_1}': self.page6_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл (өртгөөр)'), ('main_id', '=', self.id)]).mrp,
            '{page6_1_3_1}': self.page6_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл (өртгөөр)'), ('main_id', '=', self.id)]).production,
            '{page6_1_4_1}': self.page6_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл (өртгөөр)'), ('main_id', '=', self.id)]).product,
            '{page6_1_5_1}': self.page6_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл (өртгөөр)'), ('main_id', '=', self.id)]).supply_materials,

            '{page6_1_1_2}': self.page6_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн дүн'), ('main_id', '=', self.id)]).raw_materials,
            '{page6_1_2_2}': self.page6_ids.search([('name', 'ilike', 'Нэмэгдсэн дүн'), ('main_id', '=', self.id)]).mrp,
            '{page6_1_3_2}': self.page6_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн дүн'), ('main_id', '=', self.id)]).production,
            '{page6_1_4_2}': self.page6_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн дүн'), ('main_id', '=', self.id)]).product,
            '{page6_1_5_2}': self.page6_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн дүн'), ('main_id', '=', self.id)]).supply_materials,

            '{page6_1_1_3}': self.page6_ids.search(
                [('name', 'ilike', 'Хасагдсан дүн (-)'), ('main_id', '=', self.id)]).raw_materials,
            '{page6_1_2_3}': self.page6_ids.search(
                [('name', 'ilike', 'Хасагдсан дүн (-)'), ('main_id', '=', self.id)]).mrp,
            '{page6_1_3_3}': self.page6_ids.search(
                [('name', 'ilike', 'Хасагдсан дүн (-)'), ('main_id', '=', self.id)]).production,
            '{page6_1_4_3}': self.page6_ids.search(
                [('name', 'ilike', 'Хасагдсан дүн (-)'), ('main_id', '=', self.id)]).product,
            '{page6_1_5_3}': self.page6_ids.search(
                [('name', 'ilike', 'Хасагдсан дүн (-)'), ('main_id', '=', self.id)]).supply_materials,

            '{page6_1_1_4}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын гарз (-)'), ('main_id', '=', self.id)]).raw_materials,
            '{page6_1_2_4}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын гарз (-)'), ('main_id', '=', self.id)]).mrp,
            '{page6_1_3_4}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын гарз (-)'), ('main_id', '=', self.id)]).production,
            '{page6_1_4_4}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын гарз (-)'), ('main_id', '=', self.id)]).product,
            '{page6_1_5_4}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын гарз (-)'), ('main_id', '=', self.id)]).supply_materials,

            '{page6_1_1_5}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын буцаалт'), ('main_id', '=', self.id)]).raw_materials,
            '{page6_1_2_5}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын буцаалт'), ('main_id', '=', self.id)]).mrp,
            '{page6_1_3_5}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын буцаалт'), ('main_id', '=', self.id)]).production,
            '{page6_1_4_5}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын буцаалт'), ('main_id', '=', self.id)]).product,
            '{page6_1_5_5}': self.page6_ids.search(
                [('name', 'ilike', 'Үнийн бууралтын буцаалт'), ('main_id', '=', self.id)]).supply_materials,

            '{page8_1_1}': float(self.page8_ids.search(
                [('name', 'ilike', 'Урьдчилж төлсөн зардал1'), ('main_id', '=', self.id)]).first_amount),
            '{page8_1_2}': float(self.page8_ids.search(
                [('name', 'ilike', 'Урьдчилж төлсөн зардал2'), ('main_id', '=', self.id)]).last_amount),

            '{page8_2_1}': float(self.page8_ids.search(
                [('name', 'ilike', 'Урьдчилж төлсөн түрээс, даатгал1'), ('main_id', '=', self.id)]).first_amount),
            '{page8_2_2}': float(self.page8_ids.search(
                [('name', 'ilike', 'Урьдчилж төлсөн түрээс, даатгал2'), ('main_id', '=', self.id)]).last_amount),

            '{page8_3_1}': float(self.page8_ids.search([('name', 'ilike', 'Бэлтгэн нийлүүлэгчид төлсөн урьдчилгаа төлбөр1'),
                                                  ('main_id', '=', self.id)]).first_amount),
            '{page8_3_2}': float(self.page8_ids.search([('name', 'ilike', 'Бэлтгэн нийлүүлэгчид төлсөн урьдчилгаа төлбөр2'),
                                                  ('main_id', '=', self.id)]).last_amount),

            '{10_1_1}': self.page10_ids.search([('name', 'ilike', 'Өөрөө үйлдвэрлэсэн'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_2_1}': self.page10_ids.search([('name', 'ilike', 'Худалдаж авсан'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_3_1}': self.page10_ids.search([('name', 'ilike', 'Үнэ төлбөргүй авсан'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_4_1}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээний нэмэгдэл'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_5_1}': self.page10_ids.search([('name', 'ilike', 'Худалдсан'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_6_1}': self.page10_ids.search([('name', 'ilike', 'Үнэгүй шилжүүлсэн'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_7_1}': self.page10_ids.search([('name', 'ilike', 'Акталж, устгасан'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_8_1}': self.page10_ids.search([('name', 'ilike', 'Байгуулсан хорогдол'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_9_1}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр нэмэгдсэн'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_10_1}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралтын буцаалт'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_11_1}': self.page10_ids.search([('name', 'ilike', 'Данснаас хассан хөрөнгийн хорогдол'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_12_1}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр хасагдсан'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_13_1}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралт'), ('main_id', '=', self.id)], limit=1).zohiogchiin_erh,
            '{10_1_2}': self.page10_ids.search([('name', 'ilike', 'Өөрөө үйлдвэрлэсэн'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_2_2}': self.page10_ids.search([('name', 'ilike', 'Худалдаж авсан'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_3_2}': self.page10_ids.search([('name', 'ilike', 'Үнэ төлбөргүй авсан'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_4_2}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээний нэмэгдэл'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_5_2}': self.page10_ids.search([('name', 'ilike', 'Худалдсан'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_6_2}': self.page10_ids.search([('name', 'ilike', 'Үнэгүй шилжүүлсэн'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_7_2}': self.page10_ids.search([('name', 'ilike', 'Акталж, устгасан'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_8_2}': self.page10_ids.search([('name', 'ilike', 'Байгуулсан хорогдол'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_9_2}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр нэмэгдсэн'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_10_2}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралтын буцаалт'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_11_2}': self.page10_ids.search([('name', 'ilike', 'Данснаас хассан хөрөнгийн хорогдол'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_12_2}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр хасагдсан'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_13_2}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралт'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_1_3}': self.page10_ids.search([('name', 'ilike', 'Өөрөө үйлдвэрлэсэн'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_2_3}': self.page10_ids.search([('name', 'ilike', 'Худалдаж авсан'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_3_3}': self.page10_ids.search([('name', 'ilike', 'Үнэ төлбөргүй авсан'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_4_3}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээний нэмэгдэл'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_5_3}': self.page10_ids.search([('name', 'ilike', 'Худалдсан'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_6_3}': self.page10_ids.search([('name', 'ilike', 'Үнэгүй шилжүүлсэн'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_7_3}': self.page10_ids.search([('name', 'ilike', 'Акталж, устгасан'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_8_3}': self.page10_ids.search([('name', 'ilike', 'Байгуулсан хорогдол'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_9_3}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр нэмэгдсэн'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_10_3}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралтын буцаалт'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_11_3}': self.page10_ids.search([('name', 'ilike', 'Данснаас хассан хөрөнгийн хорогдол'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_12_3}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр хасагдсан'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_13_3}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралт'), ('main_id', '=', self.id)], limit=1).patent,
            '{10_1_4}': self.page10_ids.search([('name', 'ilike', 'Өөрөө үйлдвэрлэсэн'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_2_4}': self.page10_ids.search([('name', 'ilike', 'Худалдаж авсан'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_3_4}': self.page10_ids.search([('name', 'ilike', 'Үнэ төлбөргүй авсан'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_4_4}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээний нэмэгдэл'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_5_4}': self.page10_ids.search([('name', 'ilike', 'Худалдсан'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_6_4}': self.page10_ids.search([('name', 'ilike', 'Үнэгүй шилжүүлсэн'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_7_4}': self.page10_ids.search([('name', 'ilike', 'Акталж, устгасан'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_8_4}': self.page10_ids.search([('name', 'ilike', 'Байгуулсан хорогдол'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_9_4}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр нэмэгдсэн'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_10_4}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралтын буцаалт'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_11_4}': self.page10_ids.search([('name', 'ilike', 'Данснаас хассан хөрөнгийн хорогдол'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_12_4}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр хасагдсан'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_13_4}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралт'), ('main_id', '=', self.id)], limit=1).baraanii_temdeg,
            '{10_1_5}': self.page10_ids.search([('name', 'ilike', 'Өөрөө үйлдвэрлэсэн'), ('main_id', '=', self.id)], limit=1).computer,
            '{10_2_5}': self.page10_ids.search([('name', 'ilike', 'Худалдаж авсан'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_3_5}': self.page10_ids.search([('name', 'ilike', 'Үнэ төлбөргүй авсан'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_4_5}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээний нэмэгдэл'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_5_5}': self.page10_ids.search([('name', 'ilike', 'Худалдсан'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_6_5}': self.page10_ids.search([('name', 'ilike', 'Үнэгүй шилжүүлсэн'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_7_5}': self.page10_ids.search([('name', 'ilike', 'Акталж, устгасан'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_8_5}': self.page10_ids.search([('name', 'ilike', 'Байгуулсан хорогдол'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_9_5}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр нэмэгдсэн'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_10_5}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралтын буцаалт'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_11_5}': self.page10_ids.search([('name', 'ilike', 'Данснаас хассан хөрөнгийн хорогдол'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_12_5}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр хасагдсан'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_13_5}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралт'), ('main_id', '=', self.id)], limit=1).tusgai_zovshoorol,
            '{10_1_6}': self.page10_ids.search([('name', 'ilike', 'Өөрөө үйлдвэрлэсэн'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_2_6}': self.page10_ids.search([('name', 'ilike', 'Худалдаж авсан'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_3_6}': self.page10_ids.search([('name', 'ilike', 'Үнэ төлбөргүй авсан'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_4_6}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээний нэмэгдэл'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_5_6}': self.page10_ids.search([('name', 'ilike', 'Худалдсан'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_6_6}': self.page10_ids.search([('name', 'ilike', 'Үнэгүй шилжүүлсэн'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_7_6}': self.page10_ids.search([('name', 'ilike', 'Акталж, устгасан'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_8_6}': self.page10_ids.search([('name', 'ilike', 'Байгуулсан хорогдол'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_9_6}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр нэмэгдсэн'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_10_6}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралтын буцаалт'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_11_6}': self.page10_ids.search([('name', 'ilike', 'Данснаас хассан хөрөнгийн хорогдол'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_12_6}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр хасагдсан'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_13_6}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралт'), ('main_id', '=', self.id)], limit=1).gazar,
            '{10_1_7}': self.page10_ids.search([('name', 'ilike', 'Өөрөө үйлдвэрлэсэн'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_2_7}': self.page10_ids.search([('name', 'ilike', 'Худалдаж авсан'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_3_7}': self.page10_ids.search([('name', 'ilike', 'Үнэ төлбөргүй авсан'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_4_7}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээний нэмэгдэл'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_5_7}': self.page10_ids.search([('name', 'ilike', 'Худалдсан'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_6_7}': self.page10_ids.search([('name', 'ilike', 'Үнэгүй шилжүүлсэн'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_7_7}': self.page10_ids.search([('name', 'ilike', 'Акталж, устгасан'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_8_7}': self.page10_ids.search([('name', 'ilike', 'Байгуулсан хорогдол'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_9_7}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр нэмэгдсэн'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_10_7}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралтын буцаалт'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_11_7}': self.page10_ids.search([('name', 'ilike', 'Данснаас хассан хөрөнгийн хорогдол'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_12_7}': self.page10_ids.search([('name', 'ilike', 'Дахин үнэлгээгээр хасагдсан'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,
            '{10_13_7}': self.page10_ids.search([('name', 'ilike', 'Үнэ цэнийн бууралт'), ('main_id', '=', self.id)], limit=1).biet_bus_horongo,

            '{page16_2_1_1}': self.page8_ids.search(
                [('name', 'ilike', 'ААНОАТ-ын өр'), ('main_id', '=', self.id)]).first_amount,
            '{page16_2_1_2}': self.page8_ids.search(
                [('name', 'ilike', 'ААНОАТ-ын өр'), ('main_id', '=', self.id)]).last_amount,

            '{page16_2_2_1}': self.page8_ids.search(
                [('name', 'ilike', 'НӨАТ-ын өр'), ('main_id', '=', self.id)]).first_amount,
            '{page16_2_2_2}': self.page8_ids.search(
                [('name', 'ilike', 'НӨАТ-ын өр'), ('main_id', '=', self.id)]).last_amount,

            '{page16_2_3_1}': self.page8_ids.search(
                [('name', 'ilike', 'ХХОАТ-ын өр'), ('main_id', '=', self.id)]).first_amount,
            '{page16_2_3_2}': self.page8_ids.search(
                [('name', 'ilike', 'ХХОАТ-ын өр'), ('main_id', '=', self.id)]).last_amount,

            '{page16_2_4_1}': self.page8_ids.search(
                [('name', 'ilike', 'Онцгой албан татварын өр'), ('main_id', '=', self.id)]).first_amount,
            '{page16_2_4_2}': self.page8_ids.search(
                [('name', 'ilike', 'Онцгой албан татварын өр'), ('main_id', '=', self.id)]).last_amount,

            '{page16_2_5_1}': self.page8_ids.search(
                [('name', 'ilike', 'Бусад татварын өр'), ('main_id', '=', self.id)]).first_amount,
            '{page16_2_5_2}': self.page8_ids.search(
                [('name', 'ilike', 'Бусад татварын өр'), ('main_id', '=', self.id)]).last_amount,

            '{16_6_1_1}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас шууд авсан зээл'),
                                                    ('main_id', '=', self.id)]).first_amount,
            '{16_6_1_2}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас шууд авсан зээл'),
                                                    ('main_id', '=', self.id)]).first_amount_curr,
            '{16_6_1_3}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас шууд авсан зээл'),
                                                    ('main_id', '=', self.id)]).last_amount,
            '{16_6_1_4}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас шууд авсан зээл'),
                                                    ('main_id', '=', self.id)]).last_amount_curr,
            '{16_6_2_1}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас дамжуулан авсан зээл'),
                                                    ('main_id', '=', self.id)]).first_amount,
            '{16_6_2_2}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас дамжуулан авсан зээл'),
                                                    ('main_id', '=', self.id)]).first_amount_curr,
            '{16_6_2_3}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас дамжуулан авсан зээл'),
                                                    ('main_id', '=', self.id)]).last_amount,
            '{16_6_2_4}': self.page16_6_ids.search([('name', 'ilike', 'Гадаадын байгууллагаас дамжуулан авсан зээл'),
                                                    ('main_id', '=', self.id)]).last_amount_curr,
            '{16_6_3_1}': self.page16_6_ids.search([('name', 'ilike', 'Дотоодын эх үүсвэрээс авсан зээл'),
                                                    ('main_id', '=', self.id)]).first_amount,
            '{16_6_3_2}': self.page16_6_ids.search([('name', 'ilike', 'Дотоодын эх үүсвэрээс авсан зээл'),
                                                    ('main_id', '=', self.id)]).first_amount_curr,
            '{16_6_3_3}': self.page16_6_ids.search([('name', 'ilike', 'Дотоодын эх үүсвэрээс авсан зээл'),
                                                    ('main_id', '=', self.id)]).last_amount,
            '{16_6_3_4}': self.page16_6_ids.search([('name', 'ilike', 'Дотоодын эх үүсвэрээс авсан зээл'),
                                                    ('main_id', '=', self.id)]).last_amount_curr,
            '{16_6_4_1}': self.page16_6_ids.search(
                [('name', 'ilike', '(Гадаад, дотоодын зах зээлд гаргасан бонд, өрийн бичиг)'),
                 ('main_id', '=', self.id)]).first_amount,
            '{16_6_4_2}': self.page16_6_ids.search(
                [('name', 'ilike', '(Гадаад, дотоодын зах зээлд гаргасан бонд, өрийн бичиг)'),
                 ('main_id', '=', self.id)]).first_amount_curr,
            '{16_6_4_3}': self.page16_6_ids.search(
                [('name', 'ilike', '(Гадаад, дотоодын зах зээлд гаргасан бонд, өрийн бичиг)'),
                 ('main_id', '=', self.id)]).last_amount,
            '{16_6_4_4}': self.page16_6_ids.search(
                [('name', 'ilike', '(Гадаад, дотоодын зах зээлд гаргасан бонд, өрийн бичиг)'),
                 ('main_id', '=', self.id)]).last_amount_curr,

            '{17_1_1_1}': self.page17_1_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'),
                 ('main_id', '=', self.id)]).qty_1,
            '{17_1_1_2}': self.page17_1_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'),
                 ('main_id', '=', self.id)]).amount_1,
            '{17_1_1_3}': self.page17_1_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'),
                 ('main_id', '=', self.id)]).qty_2,
            '{17_1_1_4}': self.page17_1_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'),
                 ('main_id', '=', self.id)]).amount_2,
            '{17_1_2_1}': self.page17_1_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн'),
                 ('main_id', '=', self.id)]).qty_1,
            '{17_1_2_2}': self.page17_1_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн'),
                 ('main_id', '=', self.id)]).amount_1,
            '{17_1_2_3}': self.page17_1_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн'),
                 ('main_id', '=', self.id)]).qty_2,
            '{17_1_2_4}': self.page17_1_ids.search(
                [('name', 'ilike', 'Нэмэгдсэн'),
                 ('main_id', '=', self.id)]).amount_2,
            '{17_1_3_1}': self.page17_1_ids.search(
                [('name', 'ilike', 'Хасагдсан (-)'),
                 ('main_id', '=', self.id)]).qty_1,
            '{17_1_3_2}': self.page17_1_ids.search(
                [('name', 'ilike', 'Хасагдсан (-)'),
                 ('main_id', '=', self.id)]).amount_1,
            '{17_1_3_3}': self.page17_1_ids.search(
                [('name', 'ilike', 'Хасагдсан (-)'),
                 ('main_id', '=', self.id)]).qty_2,
            '{17_1_3_4}': self.page17_1_ids.search(
                [('name', 'ilike', 'Хасагдсан (-)'),
                 ('main_id', '=', self.id)]).amount_2,

            '{17_2_1_1}': self.page17_2_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'),
                 ('main_id', '=', self.id)]).undsen_hurungiin_dun,
            '{17_2_1_2}': self.page17_2_ids.search(
                [('name', 'ilike', 'Эхний үлдэгдэл'),
                 ('main_id', '=', self.id)]).biet_bus_hurungiin_dun,
            '{17_2_2_1}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлгээний нэмэгдлийн зөрүү'),
                 ('main_id', '=', self.id)])[0].undsen_hurungiin_dun,
            '{17_2_2_2}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлгээний нэмэгдлийн зөрүү'),
                 ('main_id', '=', self.id)])[0].biet_bus_hurungiin_dun,
            '{17_2_3_1}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлсэн хөрөнгийн үнэ цэнийн бууралтын гарзын буцаалт **'),
                 ('main_id', '=', self.id)]).undsen_hurungiin_dun,
            '{17_2_3_2}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлсэн хөрөнгийн үнэ цэнийн бууралтын гарзын буцаалт **'),
                 ('main_id', '=', self.id)]).biet_bus_hurungiin_dun,
            '{17_2_4_1}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлгээний нэмэгдлийн зөрүү'),
                 ('main_id', '=', self.id)])[1].undsen_hurungiin_dun,
            '{17_2_4_2}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлгээний нэмэгдлийн зөрүү'),
                 ('main_id', '=', self.id)])[1].biet_bus_hurungiin_dun,
            '{17_2_5_1}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлгээний нэмэгдлийн хэрэгжсэн дүн'),
                 ('main_id', '=', self.id)]).undsen_hurungiin_dun,
            '{17_2_5_2}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлгээний нэмэгдлийн хэрэгжсэн дүн'),
                 ('main_id', '=', self.id)]).biet_bus_hurungiin_dun,
            '{17_2_6_1}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлсэн хөрөнгийн үнэ цэнийн бууралтын гарз***'),
                 ('main_id', '=', self.id)]).undsen_hurungiin_dun,
            '{17_2_6_2}': self.page17_2_ids.search(
                [('name', 'ilike', 'Дахин үнэлсэн хөрөнгийн үнэ цэнийн бууралтын гарз***'),
                 ('main_id', '=', self.id)]).biet_bus_hurungiin_dun,

            '{17_3_1_1}': self.page17_3_ids.search(
                [('name', 'ilike', 'Гадаад үйл ажиллагааны хөрвүүлэлтээс үүссэн зөрүү'),
                 ('main_id', '=', self.id)]).initial_balance,
            '{17_3_1_2}': self.page17_3_ids.search(
                [('name', 'ilike', 'Гадаад үйл ажиллагааны хөрвүүлэлтээс үүссэн зөрүү'),
                 ('main_id', '=', self.id)]).nemegdsen,
            '{17_3_1_3}': self.page17_3_ids.search(
                [('name', 'ilike', 'Гадаад үйл ажиллагааны хөрвүүлэлтээс үүссэн зөрүү'),
                 ('main_id', '=', self.id)]).hasagdsan,
            '{17_3_2_1}': self.page17_3_ids.search(
                [('name', 'ilike', 'Бүртгэлийн валютыг толилуулгын валют руу хөрвүүлснээс үүссэн зөрүү'),
                 ('main_id', '=', self.id)]).initial_balance,
            '{17_3_2_2}': self.page17_3_ids.search(
                [('name', 'ilike', 'Бүртгэлийн валютыг толилуулгын валют руу хөрвүүлснээс үүссэн зөрүү'),
                 ('main_id', '=', self.id)]).nemegdsen,
            '{17_3_2_3}': self.page17_3_ids.search(
                [('name', 'ilike', 'Бүртгэлийн валютыг толилуулгын валют руу хөрвүүлснээс үүссэн зөрүү'),
                 ('main_id', '=', self.id)]).hasagdsan,
            '{17_3_3_1}': self.page17_3_ids.search(
                [('name', 'ilike', 'Бусад'), ('main_id', '=', self.id)]).initial_balance,
            '{17_3_3_2}': self.page17_3_ids.search([('name', 'ilike', 'Бусад'), ('main_id', '=', self.id)]).nemegdsen,
            '{17_3_3_3}': self.page17_3_ids.search([('name', 'ilike', 'Бусад'), ('main_id', '=', self.id)]).hasagdsan,

            '{22_1_1_1}': list22_1[0]['tc'], '{22_1_1_2}': list22_1[0]['hdhttc'], '{22_1_1_3}': list22_1[0]['hdhthh'],
            '{22_1_1_4}': list22_1[0]['t'],
            '{22_1_2_1}': list22_1[1]['tc'], '{22_1_2_2}': list22_1[1]['hdhttc'], '{22_1_2_3}': list22_1[1]['hdhthh'],
            '{22_1_2_4}': list22_1[1]['t'],
            '{22_1_3_1}': list22_1[2]['tc'], '{22_1_3_2}': list22_1[2]['hdhttc'], '{22_1_3_3}': list22_1[2]['hdhthh'],
            '{22_1_3_4}': list22_1[2]['t'],

            '{22_2_1_1}': list22_2[0]['ib'], '{22_2_1_2}': list22_2[0]['eb'], '{22_2_2_1}': list22_2[1]['ib'],
            '{22_2_2_2}': list22_2[1]['eb'], '{22_2_3_1}': list22_2[2]['ib'], '{22_2_3_2}': list22_2[2]['eb'],
            '{22_2_4_1}': list22_2[3]['ib'], '{22_2_4_2}': list22_2[3]['eb'], '{22_2_5_1}': list22_2[4]['ib'],
            '{22_2_5_2}': list22_2[4]['eb'],
            '{23}': self.bolzoshgui_horongo_ba_or_tolbor,
            '{24}': self.tailagnaliin_uyiin_daraah_uil_yavts,
            '{25_1_1}': investment[1 - 1]['ib'], '{25_1_2}': investment[1 - 1]['aanoh'],
            '{25_1_3}': investment[1 - 1]['uth'], '{25_1_4}': investment[1 - 1]['onth'],
            '{25_1_5}': investment[1 - 1]['bz'], '{25_1_6}': investment[1 - 1]['gsho'],
            '{25_1_7}': investment[1 - 1]['gz'], '{25_1_8}': investment[1 - 1]['gbt'],
            '{25_1_9}': investment[1 - 1]['thh'], '{25_1_10}': investment[1 - 1]['beu'],
            '{25_2_1}': investment[2 - 1]['ib'], '{25_2_2}': investment[2 - 1]['aanoh'],
            '{25_2_3}': investment[2 - 1]['uth'], '{25_2_4}': investment[2 - 1]['onth'],
            '{25_2_5}': investment[2 - 1]['bz'], '{25_2_6}': investment[2 - 1]['gsho'],
            '{25_2_7}': investment[2 - 1]['gz'], '{25_2_8}': investment[2 - 1]['gbt'],
            '{25_2_9}': investment[2 - 1]['thh'], '{25_2_10}': investment[2 - 1]['beu'],
            '{25_3_1}': investment[3 - 1]['ib'], '{25_3_2}': investment[3 - 1]['aanoh'],
            '{25_3_3}': investment[3 - 1]['uth'], '{25_3_4}': investment[3 - 1]['onth'],
            '{25_3_5}': investment[3 - 1]['bz'], '{25_3_6}': investment[3 - 1]['gsho'],
            '{25_3_7}': investment[3 - 1]['gz'], '{25_3_8}': investment[3 - 1]['gbt'],
            '{25_3_9}': investment[3 - 1]['thh'], '{25_3_10}': investment[3 - 1]['beu'],
            '{25_4_1}': investment[4 - 1]['ib'], '{25_4_2}': investment[4 - 1]['aanoh'],
            '{25_4_3}': investment[4 - 1]['uth'], '{25_4_4}': investment[4 - 1]['onth'],
            '{25_4_5}': investment[4 - 1]['bz'], '{25_4_6}': investment[4 - 1]['gsho'],
            '{25_4_7}': investment[4 - 1]['gz'], '{25_4_8}': investment[4 - 1]['gbt'],
            '{25_4_9}': investment[4 - 1]['thh'], '{25_4_10}': investment[4 - 1]['beu'],
            '{25_5_1}': investment[5 - 1]['ib'], '{25_5_2}': investment[5 - 1]['aanoh'],
            '{25_5_3}': investment[5 - 1]['uth'], '{25_5_4}': investment[5 - 1]['onth'],
            '{25_5_5}': investment[5 - 1]['bz'], '{25_5_6}': investment[5 - 1]['gsho'],
            '{25_5_7}': investment[5 - 1]['gz'], '{25_5_8}': investment[5 - 1]['gbt'],
            '{25_5_9}': investment[5 - 1]['thh'], '{25_5_10}': investment[5 - 1]['beu'],
            '{25_6_1}': investment[6 - 1]['ib'], '{25_6_2}': investment[6 - 1]['aanoh'],
            '{25_6_3}': investment[6 - 1]['uth'], '{25_6_4}': investment[6 - 1]['onth'],
            '{25_6_5}': investment[6 - 1]['bz'], '{25_6_6}': investment[6 - 1]['gsho'],
            '{25_6_7}': investment[6 - 1]['gz'], '{25_6_8}': investment[6 - 1]['gbt'],
            '{25_6_9}': investment[6 - 1]['thh'], '{25_6_10}': investment[6 - 1]['beu'],
            '{25_7_1}': investment[7 - 1]['ib'], '{25_7_2}': investment[7 - 1]['aanoh'],
            '{25_7_3}': investment[7 - 1]['uth'], '{25_7_4}': investment[7 - 1]['onth'],
            '{25_7_5}': investment[7 - 1]['bz'], '{25_7_6}': investment[7 - 1]['gsho'],
            '{25_7_7}': investment[7 - 1]['gz'], '{25_7_8}': investment[7 - 1]['gbt'],
            '{25_7_9}': investment[7 - 1]['thh'], '{25_7_10}': investment[7 - 1]['beu'],
            '{25_8_1}': investment[8 - 1]['ib'], '{25_8_2}': investment[8 - 1]['aanoh'],
            '{25_8_3}': investment[8 - 1]['uth'], '{25_8_4}': investment[8 - 1]['onth'],
            '{25_8_5}': investment[8 - 1]['bz'], '{25_8_6}': investment[8 - 1]['gsho'],
            '{25_8_7}': investment[8 - 1]['gz'], '{25_8_8}': investment[8 - 1]['gbt'],
            '{25_8_9}': investment[8 - 1]['thh'], '{25_8_10}': investment[8 - 1]['beu'],
            '{25_9_1}': investment[9 - 1]['ib'], '{25_9_2}': investment[9 - 1]['aanoh'],
            '{25_9_3}': investment[9 - 1]['uth'], '{25_9_4}': investment[9 - 1]['onth'],
            '{25_9_5}': investment[9 - 1]['bz'], '{25_9_6}': investment[9 - 1]['gsho'],
            '{25_9_7}': investment[9 - 1]['gz'], '{25_9_8}': investment[9 - 1]['gbt'],
            '{25_9_9}': investment[9 - 1]['thh'], '{25_9_10}': investment[9 - 1]['beu'],
            '{25_10_1}': investment[10 - 1]['ib'], '{25_10_2}': investment[10 - 1]['aanoh'],
            '{25_10_3}': investment[10 - 1]['uth'], '{25_10_4}': investment[10 - 1]['onth'],
            '{25_10_5}': investment[10 - 1]['bz'], '{25_10_6}': investment[10 - 1]['gsho'],
            '{25_10_7}': investment[10 - 1]['gz'], '{25_10_8}': investment[10 - 1]['gbt'],
            '{25_10_9}': investment[10 - 1]['thh'], '{25_10_10}': investment[10 - 1]['beu'],
            '{25_11_1}': investment[11 - 1]['ib'], '{25_11_2}': investment[11 - 1]['aanoh'],
            '{25_11_3}': investment[11 - 1]['uth'], '{25_11_4}': investment[11 - 1]['onth'],
            '{25_11_5}': investment[11 - 1]['bz'], '{25_11_6}': investment[11 - 1]['gsho'],
            '{25_11_7}': investment[11 - 1]['gz'], '{25_11_8}': investment[11 - 1]['gbt'],
            '{25_11_9}': investment[11 - 1]['thh'], '{25_11_10}': investment[11 - 1]['beu'],
            '{25_12_1}': investment[12 - 1]['ib'], '{25_12_2}': investment[12 - 1]['aanoh'],
            '{25_12_3}': investment[12 - 1]['uth'], '{25_12_4}': investment[12 - 1]['onth'],
            '{25_12_5}': investment[12 - 1]['bz'], '{25_12_6}': investment[12 - 1]['gsho'],
            '{25_12_7}': investment[12 - 1]['gz'], '{25_12_8}': investment[12 - 1]['gbt'],
            '{25_12_9}': investment[12 - 1]['thh'], '{25_12_10}': investment[12 - 1]['beu'],
            '{25_13_1}': investment[13 - 1]['ib'], '{25_13_2}': investment[13 - 1]['aanoh'],
            '{25_13_3}': investment[13 - 1]['uth'], '{25_13_4}': investment[13 - 1]['onth'],
            '{25_13_5}': investment[13 - 1]['bz'], '{25_13_6}': investment[13 - 1]['gsho'],
            '{25_13_7}': investment[13 - 1]['gz'], '{25_13_8}': investment[13 - 1]['gbt'],
            '{25_13_9}': investment[13 - 1]['thh'], '{25_13_10}': investment[13 - 1]['beu'],
            '{25_14_1}': investment[14 - 1]['ib'], '{25_14_2}': investment[14 - 1]['aanoh'],
            '{25_14_3}': investment[14 - 1]['uth'], '{25_14_4}': investment[14 - 1]['onth'],
            '{25_14_5}': investment[14 - 1]['bz'], '{25_14_6}': investment[14 - 1]['gsho'],
            '{25_14_7}': investment[14 - 1]['gz'], '{25_14_8}': investment[14 - 1]['gbt'],
            '{25_14_9}': investment[14 - 1]['thh'], '{25_14_10}': investment[14 - 1]['beu'],
            '{25_15_1}': investment[15 - 1]['ib'], '{25_15_2}': investment[15 - 1]['aanoh'],
            '{25_15_3}': investment[15 - 1]['uth'], '{25_15_4}': investment[15 - 1]['onth'],
            '{25_15_5}': investment[15 - 1]['bz'], '{25_15_6}': investment[15 - 1]['gsho'],
            '{25_15_7}': investment[15 - 1]['gz'], '{25_15_8}': investment[15 - 1]['gbt'],
            '{25_15_9}': investment[15 - 1]['thh'], '{25_15_10}': investment[15 - 1]['beu'],
            '{25_16_1}': investment[16 - 1]['ib'], '{25_16_2}': investment[16 - 1]['aanoh'],
            '{25_16_3}': investment[16 - 1]['uth'], '{25_16_4}': investment[16 - 1]['onth'],
            '{25_16_5}': investment[16 - 1]['bz'], '{25_16_6}': investment[16 - 1]['gsho'],
            '{25_16_7}': investment[16 - 1]['gz'], '{25_16_8}': investment[16 - 1]['gbt'],
            '{25_16_9}': investment[16 - 1]['thh'], '{25_16_10}': investment[16 - 1]['beu'],
            '{25_17_1}': investment[17 - 1]['ib'], '{25_17_2}': investment[17 - 1]['aanoh'],
            '{25_17_3}': investment[17 - 1]['uth'], '{25_17_4}': investment[17 - 1]['onth'],
            '{25_17_5}': investment[17 - 1]['bz'], '{25_17_6}': investment[17 - 1]['gsho'],
            '{25_17_7}': investment[17 - 1]['gz'], '{25_17_8}': investment[17 - 1]['gbt'],
            '{25_17_9}': investment[17 - 1]['thh'], '{25_17_10}': investment[17 - 1]['beu'],
            '{25_18_1}': investment[18 - 1]['ib'], '{25_18_2}': investment[18 - 1]['aanoh'],
            '{25_18_3}': investment[18 - 1]['uth'], '{25_18_4}': investment[18 - 1]['onth'],
            '{25_18_5}': investment[18 - 1]['bz'], '{25_18_6}': investment[18 - 1]['gsho'],
            '{25_18_7}': investment[18 - 1]['gz'], '{25_18_8}': investment[18 - 1]['gbt'],
            '{25_18_9}': investment[18 - 1]['thh'], '{25_18_10}': investment[18 - 1]['beu'],
            '{25_19_1}': investment[19 - 1]['ib'], '{25_19_2}': investment[19 - 1]['aanoh'],
            '{25_19_3}': investment[19 - 1]['uth'], '{25_19_4}': investment[19 - 1]['onth'],
            '{25_19_5}': investment[19 - 1]['bz'], '{25_19_6}': investment[19 - 1]['gsho'],
            '{25_19_7}': investment[19 - 1]['gz'], '{25_19_8}': investment[19 - 1]['gbt'],
            '{25_19_9}': investment[19 - 1]['thh'], '{25_19_10}': investment[19 - 1]['beu'],
            '{25_20_1}': investment[20 - 1]['ib'], '{25_20_2}': investment[20 - 1]['aanoh'],
            '{25_20_3}': investment[20 - 1]['uth'], '{25_20_4}': investment[20 - 1]['onth'],
            '{25_20_5}': investment[20 - 1]['bz'], '{25_20_6}': investment[20 - 1]['gsho'],
            '{25_20_7}': investment[20 - 1]['gz'], '{25_20_8}': investment[20 - 1]['gbt'],
            '{25_20_9}': investment[20 - 1]['thh'], '{25_20_10}': investment[20 - 1]['beu'],
            '{25_21_1}': investment[21 - 1]['ib'], '{25_21_2}': investment[21 - 1]['aanoh'],
            '{25_21_3}': investment[21 - 1]['uth'], '{25_21_4}': investment[21 - 1]['onth'],
            '{25_21_5}': investment[21 - 1]['bz'], '{25_21_6}': investment[21 - 1]['gsho'],
            '{25_21_7}': investment[21 - 1]['gz'], '{25_21_8}': investment[21 - 1]['gbt'],
            '{25_21_9}': investment[21 - 1]['thh'], '{25_21_10}': investment[21 - 1]['beu'],
            '{25_22_1}': investment[22 - 1]['ib'], '{25_22_2}': investment[22 - 1]['aanoh'],
            '{25_22_3}': investment[22 - 1]['uth'], '{25_22_4}': investment[22 - 1]['onth'],
            '{25_22_5}': investment[22 - 1]['bz'], '{25_22_6}': investment[22 - 1]['gsho'],
            '{25_22_7}': investment[22 - 1]['gz'], '{25_22_8}': investment[22 - 1]['gbt'],
            '{25_22_9}': investment[22 - 1]['thh'], '{25_22_10}': investment[22 - 1]['beu'],
            '{25_23_1}': investment[23 - 1]['ib'], '{25_23_2}': investment[23 - 1]['aanoh'],
            '{25_23_3}': investment[23 - 1]['uth'], '{25_23_4}': investment[23 - 1]['onth'],
            '{25_23_5}': investment[23 - 1]['bz'], '{25_23_6}': investment[23 - 1]['gsho'],
            '{25_23_7}': investment[23 - 1]['gz'], '{25_23_8}': investment[23 - 1]['gbt'],
            '{25_23_9}': investment[23 - 1]['thh'], '{25_23_10}': investment[23 - 1]['beu'],
        }
        sheet = wb.active
        rows = sheet.rows
        for row in rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border
        sheet23_24 = wb['23-24']
        for row in sheet23_24.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border
        sheet25 = wb['25']
        for row in sheet25.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border

        sheet = wb['1-2']
        row = 9
        for page in self.page1_1_ids:
            sheet.cell(row=row, column=3, value='%s' % page.name)
            row += 1
        row += 4
        for page in self.page1_2_ids:
            sheet.cell(row=row, column=3, value='%s' % page.name)
            row += 1
        row += 4
        for page in self.page1_3_ids:
            sheet.cell(row=row, column=3, value='%s' % page.name)
            row += 1

        row = 51
        number = 1
        sheet = wb['20-22']
        for page in self.page22_3_ids:
            sheet.cell(row=row, column=1, value='%s' % number)
            sheet.cell(row=row, column=2, value='%s' % page.name)
            sheet.cell(row=row, column=4, value='%s' % page.transaction_description)
            sheet.cell(row=row, column=6, value='%s' % page.amount)
            sheet.cell(row=row, column=8, value='%s' % page.description)
            number += 1
            row += 1
        for row in sheet.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border
                   
        sheet = wb['3-4']
        for row in sheet.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border
                    
        sheet = wb['5-8']
        for row in sheet.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border
        
        sheet = wb['16-17']
        for row in sheet.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border
         
        row = 4
        number = 1
        for page in self.page16_5_ids:
            sheet.cell(row=row, column=1, value='%s' % number).alignment = Alignment(horizontal='center')
            number += 1
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            sheet.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)

            sheet.cell(row=row, column=2, value='%s' % page.name).alignment = Alignment(horizontal='left')
            sheet.cell(row=row, column=5, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=7, value=page.last_amount).alignment = Alignment(horizontal='right')
            row += 1
            if number > 2 and number != len(self.page16_5_ids):
                sheet.insert_rows(row, amount=1)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        sheet.cell(row=row, column=2, value=u'Нийт дүн').alignment = Alignment(horizontal='center')
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        sheet.cell(row=row, column=5, value='=SUM(E4:F%s)' % str(2 + number)).alignment = Alignment(horizontal='right')
        sheet.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        sheet.cell(row=row, column=7, value='=SUM(G4:H%s)' % str(2 + number)).alignment = Alignment(horizontal='right')
        number16_5 = number

        row += number + 10
        number = 2
        for page in self.page16_6_ids:
            if page.name not in (
                    'Гадаадын байгууллагаас шууд авсан зээл',
                    'Гадаадын байгууллагаас дамжуулан авсан зээл',
                    'Дотоодын эх үүсвэрээс авсан зээл',
                    '(Гадаад, дотоодын зах зээлд гаргасан бонд, өрийн бичиг)',
                    'Урт хугацаат зээлийн дүн',
                    'Бусад урт хугацаат өр төлбөрийн дүн'):
                if number == 4:
                    sheet.unmerge_cells('A26:H26')
                if number > 2:
                    sheet.cell(row=row, column=1, value='2.%s' % str(number)).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=2, value='%s' % page.name).alignment = Alignment(horizontal='left')
                sheet.cell(row=row, column=5, value=page.first_amount).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=6, value=page.first_amount_curr).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=7, value=page.last_amount).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=8, value=page.last_amount_curr).alignment = Alignment(horizontal='center')
                row += 1
                number += 1
                if number != 2:
                    if number == 3:
                        sheet.insert_rows(row + 1, amount=1)
                    else:
                        sheet.insert_rows(row, amount=1)
            elif page.name == 'Урт хугацаат зээлийн дүн':
                sheet.cell(row=number16_5 + 12, column=5, value='=SUM(E%s:E%s)' % (number16_5 + 13, number16_5 + 15)).alignment = Alignment(horizontal='center')
#                 sheet.cell(row=number16_5 + 12, column=6, value='=SUM(F%s:F%s)' % (number16_5 + 13, number16_5 + 15)).alignment = Alignment(horizontal='center')
#                 sheet.cell(row=number16_5 + 12, column=7, value='=SUM(G%s:G%s)' % (number16_5 + 13, number16_5 + 15)).alignment = Alignment(horizontal='center')
#                 sheet.cell(row=number16_5 + 12, column=8, value='=SUM(H%s:H%s)' % (number16_5 + 13, number16_5 + 15)).alignment = Alignment(horizontal='center')
            elif page.name == 'Бусад урт хугацаат өр төлбөрийн дүн':
                sheet.cell(row=number16_5 + 16, column=5, value='=SUM(E%s:E%s)' % (number16_5 + 17, len(self.page16_6_ids) + number16_5 + 11)).alignment = Alignment(horizontal='center')
                sheet.cell(row=number16_5 + 16, column=6, value='=SUM(F%s:F%s)' % (number16_5 + 17, len(self.page16_6_ids) + number16_5 + 11)).alignment = Alignment(horizontal='center')
                sheet.cell(row=number16_5 + 16, column=7, value='=SUM(G%s:G%s)' % (number16_5 + 17, len(self.page16_6_ids) + number16_5 + 11)).alignment = Alignment(horizontal='center')
                sheet.cell(row=number16_5 + 16, column=8, value='=SUM(H%s:H%s)' % (number16_5 + 17, len(self.page16_6_ids) + number16_5 + 11)).alignment = Alignment(horizontal='center')
        for row in sheet.rows:
            for cell in row:

                if cell.value == '=SUM(F45:F47)':
                    cell.value = '=SUM(F%s:F%s)' % (
                        45 + number16_5 - 3 + number - 2,
                        47 + number16_5 - 3 + number - 2)
                if cell.value == '=SUM(G45:G47)':
                    cell.value = '=SUM(G%s:G%s)' % (
                        45 + number16_5 - 3 + number - 2,
                        47 + number16_5 - 3 + number - 2)
                if cell.value == '=SUM(F42:F43)':
                    cell.value = '=SUM(F%s:F%s)' % (
                        42 + number16_5 - 3 + number - 2,
                        43 + number16_5 - 3 + number - 2)
                if cell.value == '=SUM(G42:G43)':
                    cell.value = '=SUM(G%s:G%s)' % (
                        42 + number16_5 - 3 + number - 2,
                        43 + number16_5 - 3 + number - 2)
                if cell.value == '=+F40+F41-F44':
                    cell.value = '=+F%s+F%s-F%s' % (
                        40 + number16_5 - 3 + number - 2,
                        41 + number16_5 - 3 + number - 2,
                        47 + number16_5 - 3 + number - 2)
                if cell.value == '=+G40+G41-G44':
                    cell.value = '=+G%s+G%s-G%s' % (
                        40 + number16_5 - 3 + number - 2,
                        41 + number16_5 - 3 + number - 2,
                        47 + number16_5 - 3 + number - 2)
                if cell.value == '=+F40+G40':
                    sheet.unmerge_cells('B50:H51')
                    sheet.unmerge_cells('B53:H55')
                    cell.value = '=+F%s + G%s' % (
                        40 + number16_5 - 3 + number - 2,
                        40 + number16_5 - 3 + number - 2)
                if cell.value == '=+F41+G41':
                    cell.value = '=+F%s + G%s' % (
                        41 + number16_5 - 3 + number - 2,
                        41 + number16_5 - 3 + number - 2)
                if cell.value == '=+F42+G42':
                    cell.value = '=+F%s + G%s' % (
                        42 + number16_5 - 3 + number - 2,
                        42 + number16_5 - 3 + number - 2)
                if cell.value == '=+F43+G43':
                    cell.value = '=+F%s + G%s' % (
                        43 + number16_5 - 3 + number - 2,
                        43 + number16_5 - 3 + number - 2)
                if cell.value == '=+F44+G44':
                    cell.value = '=+F%s + G%s' % (
                        44 + number16_5 - 3 + number - 2,
                        44 + number16_5 - 3 + number - 2)
                if cell.value == '=+F45+G45':
                    cell.value = '=+F%s + G%s' % (
                        45 + number16_5 - 3 + number - 2,
                        45 + number16_5 - 3 + number - 2)
                if cell.value == '=+F46+G46':
                    cell.value = '=+F%s + G%s' % (
                        46 + number16_5 - 3 + number - 2,
                        46 + number16_5 - 3 + number - 2)
                if cell.value == '=+F47+G47':
                    cell.value = '=+F%s + G%s' % (
                        47 + number16_5 - 3 + number - 2,
                        47 + number16_5 - 3 + number - 2)
                if cell.value == '=+F48+G48':
                    cell.value = '=+F%s + G%s' % (
                        48 + number16_5 - 3 + number - 2,
                        48 + number16_5 - 3 + number - 2)
                if cell.value == '=+E32+G32':
                    cell.value = '=+E%s + G%s' % (
                        32 + number16_5 - 3 + number - 2,
                        32 + number16_5 - 3 + number - 2)
                if cell.value == '=+E33+G33':
                    cell.value = '=+E%s + G%s' % (
                        33 + number16_5 - 3 + number - 2,
                        33 + number16_5 - 3 + number - 2)
                if cell.value == '=+E34+G34':
                    cell.value = '=+E%s + G%s' % (
                        34 + number16_5 - 3 + number - 2,
                        34 + number16_5 - 3 + number - 2)
                if cell.value == '=+D32+D33-D34':
                    sheet.unmerge_cells('B39:E39')
                    sheet.unmerge_cells('B40:E40')
                    cell.value = '=+D%s + D%s - D%s' % (
                        32 + number16_5 - 3 + number - 2,
                        33 + number16_5 - 3 + number - 2,
                        34 + number16_5 - 3 + number - 2)
                if cell.value == '=+E32+E33-E34':
                    cell.value = '=+E%s + E%s - E%s' % (
                        32 + number16_5 - 3 + number - 2,
                        33 + number16_5 - 3 + number - 2,
                        34 + number16_5 - 3 + number - 2)
                if cell.value == '=+F32+F33-F34':
                    cell.value = '=+F%s + F%s - F%s' % (
                        32 + number16_5 - 3 + number - 2,
                        33 + number16_5 - 3 + number - 2,
                        34 + number16_5 - 3 + number - 2)
                if cell.value == '=+G32+G33-G34':
                    cell.value = '=+G%s + G%s - G%s' % (
                        32 + number16_5 - 3 + number - 2,
                        33 + number16_5 - 3 + number - 2,
                        34 + number16_5 - 3 + number - 2)
                if cell.value == '=+H32+H33-H34':
                    cell.value = '=+H%s + H%s - H%s' % (
                        32 + number16_5 - 3 + number - 2,
                        33 + number16_5 - 3 + number - 2,
                        34 + number16_5 - 3 + number - 2)

        # 9
        sheet = wb['9']
        row = 6
        number = 0
        down = 0
        for page in self.page9_1_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 8 else None
            down += 1 if number > 8 else 0
            if number == 5:
                sheet.cell(row=row, column=3, value='=SUM(C12:C%s)' % str(7 + len(self.page9_1_ids)))
                sheet.cell(row=row, column=4, value='=SUM(D12:D%s)' % str(7 + len(self.page9_1_ids)))
                sheet.cell(row=row, column=5, value='=SUM(E12:E%s)' % str(7 + len(self.page9_1_ids)))
                sheet.cell(row=row, column=6, value='=SUM(F12:F%s)' % str(7 + len(self.page9_1_ids)))
                sheet.cell(row=row, column=7, value='=SUM(G12:G%s)' % str(7 + len(self.page9_1_ids)))
                sheet.cell(row=row, column=8, value='=SUM(H12:H%s)' % str(7 + len(self.page9_1_ids)))
                sheet.cell(row=row, column=9, value='=SUM(I12:I%s)' % str(7 + len(self.page9_1_ids)))
                row += 1
            sheet.cell(row=row, column=2, value=page.name)
#             sheet.cell(row=row, column=3, value=int(page.gazriin_saijruulalt))
#             sheet.cell(row=row, column=4, value=float(page.barilga))
#             sheet.cell(row=row, column=5, value=float(page.mashin_totu))
#             sheet.cell(row=row, column=6, value=float(page.teevriin))
#             sheet.cell(row=row, column=7, value=float(page.tavilga))
#             sheet.cell(row=row, column=8, value=float(page.computer))
#             sheet.cell(row=row, column=9, value=float(page.busad))

        number = 0
        row += 1 if len(self.page9_1_ids) == 7 else 0
        for page in self.page9_2_ids:
            row += 1
            number += 1
            row += 4 if number == 3 else 0
            row += 1 if number == 6 else 0
            sheet.cell(row=row, column=2, value=page.name)
            sheet.cell(row=row, column=3, value=(page.gazriin_saijruulalt))
            sheet.cell(row=row, column=4, value=(page.barilga))
            sheet.cell(row=row, column=5, value=(page.mashin_totu))
            sheet.cell(row=row, column=6, value=(page.teevriin))
            sheet.cell(row=row, column=7, value=(page.tavilga))
            sheet.cell(row=row, column=8, value=(page.computer))
            sheet.cell(row=row, column=9, value=(page.busad))

#             sheet.cell(row=row, column=6, value=page.first_amount).alignment = Alignment(horizontal='right')

        sheet.unmerge_cells('A34:J36')
        row += 6
        sheet.merge_cells('A%s:J%s' % (row, row + 2))
        row += 3
        sheet.merge_cells('A%s:J%s' % (row, row + 2))
        sheet.cell(row=row, column=1, value=self.page9_describe)

        # 10
        sheet = wb['10']
        for row in sheet.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
        row = 35
        sheet.merge_cells('A%s:H%s' % (row, row + 2))
        sheet.cell(row=row, column=1, value=self.page10_describe)

        # 11-14
        sheet = wb['11-14']
        sheet.unmerge_cells('A38:H43')
        # 11
        down = 0
        row = 3
        sheet.cell(row=row + 1, column=2, value='')
        sheet.cell(row=row + 1, column=3, value='')
        sheet.cell(row=row + 1, column=6, value='')
        sheet.cell(row=row + 1, column=7, value='')
        sheet.cell(row=row + 1, column=8, value='')
        number = 0
        for page in self.page11_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 2 else None
            down += 1 if number > 2 else 0
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.merge_cells('C%s:E%s' % (row, row))
            sheet.cell(row=row, column=3, value=page.ehelsen_ognoo).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=6, value=page.duusgalin_huvi).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=7, value=page.niit_tosovt_ortog).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=8, value=page.etssiin_hugatsaa).alignment = Alignment(horizontal='center')
        if number < 2:
            row += 2 - number
            number += 2 - number
        row += 1
        number += 1
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=7, value='=sum(G4:G%s)' % str(2 + number)).alignment = Alignment(horizontal='right')
        down_sum = down

        # 12
        row += 1
        sheet.unmerge_cells('A8:H8')
        row += 1
        sheet.merge_cells('A%s:H%s' % (row, row))
        row += 2
        sheet.merge_cells('A%s:A%s' % (row, row + 1))
        sheet.merge_cells('B%s:B%s' % (row, row + 1))
        sheet.merge_cells('C%s:C%s' % (row, row + 1))
        sheet.merge_cells('D%s:D%s' % (row, row + 1))
        sheet.merge_cells('E%s:F%s' % (row, row))
        sheet.merge_cells('G%s:H%s' % (row, row))
        row += 1
        sheet.cell(row=row + 1, column=2, value='')
        sheet.cell(row=row + 1, column=3, value='')
        sheet.cell(row=row + 1, column=4, value='')
        sheet.cell(row=row + 1, column=5, value='')
        sheet.cell(row=row + 1, column=6, value='')
        sheet.cell(row=row + 1, column=7, value='')
        sheet.cell(row=row + 1, column=8, value='')
        number = 0
        for page in self.page12_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 2 else None
            down += 1 if number > 2 else 0
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.cell(row=row, column=3, value=page.age).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=4, value=page.gender).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=5, value=page.first_qty).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=6, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=7, value=page.last_qty).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=8, value=page.last_amount).alignment = Alignment(horizontal='right')
        if number < 2:
            row += 2 - number
            number += 2 - number
        row += 1
        number += 1
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=6, value='=SUM(F%s:F%s)' % (down_sum + 12, down_sum + 10 + number)).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=8, value='=SUM(H%s:H%s)' % (down_sum + 12, down_sum + 10 + number)).alignment = Alignment(horizontal='right')
        row += 2
        sheet.merge_cells('A%s:H%s' % (row, row))
        row += 1
        sheet.merge_cells('A%s:H%s' % (row, row + 1))
        sheet.cell(row=row, column=1, value=self.page12_describe).alignment = Alignment(horizontal='left')
        down_sum = down

        # 13
        sheet.unmerge_cells('A20:H20')
        sheet.unmerge_cells('A29:H31')
        row += 3
        sheet.merge_cells('A%s:H%s' % (row, row))
        row += 2
        sheet.merge_cells('A%s:A%s' % (row, row + 1))
        sheet.merge_cells('B%s:B%s' % (row, row + 1))
        sheet.merge_cells('C%s:F%s' % (row, row))
        sheet.merge_cells('G%s:H%s' % (row, row))
        row += 1
        sheet.merge_cells('C%s:D%s' % (row, row))
        sheet.merge_cells('E%s:F%s' % (row, row))
        sheet.cell(row=row + 1, column=2, value='')
        sheet.cell(row=row + 1, column=3, value='')
        sheet.cell(row=row + 1, column=5, value='')
        sheet.cell(row=row + 1, column=7, value='')
        sheet.cell(row=row + 1, column=8, value='')
        sheet.cell(row=row + 2, column=2, value='')
        sheet.cell(row=row + 2, column=3, value='')
        sheet.cell(row=row + 2, column=5, value='')
        sheet.cell(row=row + 2, column=7, value='')
        sheet.cell(row=row + 2, column=8, value='')
        number = 0
        for page in self.page13_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 3 else None
            down += 1 if number > 3 else 0
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.merge_cells('C%s:D%s' % (row, row))
            sheet.cell(row=row, column=3, value=page.first_percent).alignment = Alignment(horizontal='right')
            sheet.merge_cells('E%s:F%s' % (row, row))
            sheet.cell(row=row, column=5, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=7, value=page.last_percent).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=8, value=page.last_amount).alignment = Alignment(horizontal='right')
        if number < 3:
            row += 3 - number
            number += 3 - number
        row += 1
        number += 1
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        sheet.merge_cells('E%s:F%s' % (row, row))
        sheet.cell(row=row, column=5, value='=sum(E%s:F%s)' % (str(24 + down_sum), str(22 + down_sum + number))).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=8, value='=sum(H%s:H%s)' % (str(24 + down_sum), str(22 + down_sum + number))).alignment = Alignment(horizontal='right')
        row += 2
        sheet.merge_cells('A%s:H%s' % (row, row + 2))
        row += 3
        sheet.merge_cells('A%s:H%s' % (row, row + 2))
        sheet.cell(row=row, column=1, value=self.page13_describe)

        # 14
        sheet.unmerge_cells('A36:H36')
        row += 4
        sheet.merge_cells('A%s:H%s' % (row, row))
        row += 2
        sheet.merge_cells('A%s:H%s' % (row, row + 5))
        row += 6
        sheet.merge_cells('A%s:H%s' % (row, row + 4))
        sheet.cell(row=row, column=1, value=self.page14_describe)

        sheet = wb['15-16']
        for row in sheet.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border
        # 15
        sheet.unmerge_cells('A8:G8')
        row = 3
        number = 0
        down = 0
        for page in self.page15_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 2 else None
            down += 1 if number > 2 else 0
            sheet.merge_cells('B%s:C%s' % (row, row))
            sheet.merge_cells('D%s:E%s' % (row, row))
            sheet.merge_cells('F%s:G%s' % (row, row))
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.cell(row=row, column=4, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=6, value=page.last_amount).alignment = Alignment(horizontal='right')
        if number < 2:
            row += 2 - number
            number += 2 - number
        row += 1
        number += 1
        sheet.merge_cells('B%s:C%s' % (row, row))
        sheet.merge_cells('D%s:E%s' % (row, row))
        sheet.cell(row=row, column=4, value='=SUM(D4:E%s)' % str(number + 2)).alignment = Alignment(horizontal='right')
        sheet.merge_cells('F%s:G%s' % (row, row))
        sheet.cell(row=row, column=6, value='=SUM(F4:G%s)' % str(number + 2)).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        down_sum = down

        # 16.1
        row += 2
        sheet.merge_cells('A%s:G%s' % (row, row))
        row += 2
        sheet.merge_cells('A%s:G%s' % (row, row))
        row += 2
        number = 0
        for page in self.page16_1_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 3 else None
            down += 1 if number > 3 else 0
            sheet.merge_cells('B%s:C%s' % (row, row))
            sheet.merge_cells('D%s:E%s' % (row, row))
            sheet.merge_cells('F%s:G%s' % (row, row))
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.cell(row=row, column=4, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=6, value=page.last_amount).alignment = Alignment(horizontal='right')
        if number < 3:
            row += 3 - number
            number += 3 - number
        row += 1
        number += 1
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        sheet.merge_cells('B%s:C%s' % (row, row))
        sheet.merge_cells('D%s:E%s' % (row, row))
        sheet.cell(row=row, column=4, value='=SUM(D%s:E%s)' % (str(13 + down_sum), str(11 + down_sum + number))).alignment = Alignment(horizontal='right')
        sheet.merge_cells('F%s:G%s' % (row, row))
        sheet.cell(row=row, column=6, value='=SUM(F%s:G%s)' % (str(13 + down_sum), str(11 + down_sum + number))).alignment = Alignment(horizontal='right')
        down_sum = down

        # 16.2
        row += 2
        sheet.merge_cells('A%s:G%s' % (row, row))
        row += 2
        number = 0
        for page in self.page16_2_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 6 else None
            down += 1 if number > 6 else 0
            sheet.merge_cells('B%s:C%s' % (row, row))
            sheet.merge_cells('D%s:E%s' % (row, row))
            sheet.merge_cells('F%s:G%s' % (row, row))
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.cell(row=row, column=4, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=6, value=page.last_amount).alignment = Alignment(horizontal='right')
        if number == 5:
            row += 1
            number += 1
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.merge_cells('B%s:C%s' % (row, row))
            sheet.merge_cells('D%s:E%s' % (row, row))
            sheet.cell(row=row, column=4, value='').alignment = Alignment(horizontal='right')
            sheet.merge_cells('F%s:G%s' % (row, row))
            sheet.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        if number < 6:
            row += 6 - number
            number += 6 - number
        row += 1
        number += 1
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        sheet.merge_cells('B%s:C%s' % (row, row))
        sheet.merge_cells('D%s:E%s' % (row, row))
        sheet.cell(row=row, column=4, value='=SUM(D%s:E%s)' % (str(21 + down_sum), str(19 + down_sum + number))).alignment = Alignment(horizontal='right')
        sheet.merge_cells('F%s:G%s' % (row, row))
        sheet.cell(row=row, column=6, value='=SUM(F%s:G%s)' % (str(21 + down_sum), str(19 + down_sum + number))).alignment = Alignment(horizontal='right')
        down_sum = down

        # 16.3
        sheet.unmerge_cells('A31:A32')
        sheet.unmerge_cells('B31:C32')
        row += 2
        sheet.merge_cells('A%s:G%s' % (row, row))
        row += 2
        sheet.merge_cells('A%s:A%s' % (row, row + 1))
        sheet.merge_cells('B%s:C%s' % (row, row + 1))
        sheet.merge_cells('D%s:E%s' % (row, row))
        sheet.merge_cells('F%s:G%s' % (row, row))
        row += 1
        number = 0
        for page in self.page16_3_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 3 else None
            down += 1 if number > 3 else 0
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.merge_cells('B%s:C%s' % (row, row))
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.cell(row=row, column=4, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=5, value=page.first_amount_curr).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=6, value=page.last_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=7, value=page.last_amount_curr).alignment = Alignment(horizontal='right')
        # if number == 2:
        #     row += 1
        #     number += 1
        #     sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        #     sheet.merge_cells('B%s:C%s' % (row, row))
        if number < 3:
            row += 3 - number
            number += 3 - number
        row += 1
        number += 1
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        sheet.merge_cells('B%s:C%s' % (row, row))
        sheet.cell(row=row, column=4, value='=SUM(D%s:D%s)' % (str(33 + down_sum), str(31 + down_sum + number))).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=5, value='=SUM(E%s:E%s)' % (str(33 + down_sum), str(31 + down_sum + number))).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=6, value='=SUM(F%s:F%s)' % (str(33 + down_sum), str(31 + down_sum + number))).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=7, value='=SUM(G%s:G%s)' % (str(33 + down_sum), str(31 + down_sum + number))).alignment = Alignment(horizontal='right')
        down_sum = down

        # 16.4
        row += 2
        sheet.merge_cells('A%s:G%s' % (row, row))
        row += 2
        number = 0
        for page in self.page16_4_ids:
            row += 1
            number += 1
            sheet.insert_rows(row, amount=1) if number > 3 else None
            down += 1 if number > 3 else 0
            sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
            sheet.cell(row=row, column=2, value=page.name).alignment = Alignment(horizontal='left')
            sheet.cell(row=row, column=3, value=page.first_amount).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=4, value=page.add).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=5, value=page.sub).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=6, value=page.useless).alignment = Alignment(horizontal='right')
            sheet.cell(row=row, column=7, value='=+C%s + D%s - E%s - F%s' % (row, row, row, row)).alignment = Alignment(horizontal='right')
        if number == 2:
            row += 1
            number += 1
            sheet.cell(row=row, column=7, value='=+C%s + D%s - E%s - F%s' % (row, row, row, row)).alignment = Alignment(horizontal='right')
        row += 1
        number += 1
        sheet.cell(row=row, column=1, value=number).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        sheet.cell(row=row, column=3, value='=SUM(C%s:C%s)' % (41 + down_sum, 39 + down_sum + number)).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=4, value='=SUM(D%s:D%s)' % (41 + down_sum, 39 + down_sum + number)).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=5, value='=SUM(E%s:E%s)' % (41 + down_sum, 39 + down_sum + number)).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=6, value='=SUM(F%s:F%s)' % (41 + down_sum, 39 + down_sum + number)).alignment = Alignment(horizontal='right')
        sheet.cell(row=row, column=7, value='=SUM(G%s:G%s)' % (41 + down_sum, 39 + down_sum + number)).alignment = Alignment(horizontal='right')

        sheet = wb['17-19']
        for row in sheet.rows:
            for cell in row:
                if cell.value in dicts:
                    cell.value = dicts[cell.value]
                    cell.border = thin_border

        from io import BytesIO
        buffer = BytesIO()

        wb.save(buffer)

        filename = u"Санхүүгийн тайлан тохиргоо_%s.xlsx" % (time.strftime('%Y%m%d_%H%M'),)
#         report_excel_output_obj = self.env['oderp.report.excel.output'].with_context(filename_prefix=
#                                                                                      u'Санхүүгийн тайлан тохиргоо',
#                                                                                      form_title=file_name).create({})
#         image = open(generate_file, 'rb')
#         image_read = image.read()
#         report_excel_output_obj.filedata = base64.encodestring(image_read)
#         return report_excel_output_obj.export_report()
#         book.save(buffer)
        buffer.seek(0)
        
#         filename = "balance_%s.xls" % (time.strftime('%Y%m%d_%H%M'),)
        out = base64.encodebytes(buffer.getvalue())
        buffer.close()
        
        excel_id = self.env['report.excel.output'].create({
                                'data':out,
                                'name':filename
        })
        return {
             'type' : 'ir.actions.act_url',
             'url': "web/content/?model=report.excel.output&id=" + str(excel_id.id) + "&filename_field=filename&download=true&field=data&filename=" + excel_id.name,
             'target': 'new',
        }    
                


class AccountDisclosureReportMainPage1_1(models.Model):
    _name = "account.disclosure.main.page1_1"
    _description = u"Үндсэн үйл ажиллагааны чиглэл /төрөл/"

    name = fields.Text(u'Үндсэн үйл ажиллагааны чиглэл /төрөл/')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage1_2(models.Model):
    _name = "account.disclosure.main.page1_2"
    _description = u"Туслах үйл ажиллагааны чиглэл /төрөл/"

    name = fields.Text(u'Туслах үйл ажиллагааны чиглэл /төрөл/')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage1_3(models.Model):
    _name = "account.disclosure.main.page1_3"
    _description = u"Салбар, төлөөлөгчийн газрын нэр, байршил"

    name = fields.Text(u'Салбар, төлөөлөгчийн газрын нэр, байршил')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage3(models.Model):
    _name = "account.disclosure.main.page3"
    _description = u"3. МӨНГӨ БА ТҮҮНТЭЙ АДИЛТГАХ ХӨРӨНГӨ"

    name = fields.Char(u'Үзүүлэлт')
    income = fields.Float(u'Эхний үлдэгдэл')
    expense = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage4_1(models.Model):
    _name = "account.disclosure.main.page4_1"
    _description = u"4.1 Дансны авлага"

    name = fields.Char(u'Үзүүлэлт')
    income = fields.Float(u'Дансны авлага')
    expense = fields.Float(u'Найдваргүй авлагын хасагдуулга')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage4_2(models.Model):
    _name = "account.disclosure.main.page4_2"
    _description = u"4.2 Татвар, нийгмийн даатгалын шимтгэл (НДШ)-ийн  авлага"

    name = fields.Char(u'Үзүүлэлт')
    income = fields.Float(u'Эхний үлдэгдэл')
    expense = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage4_3(models.Model):
    _name = "account.disclosure.main.page4_3"
    _description = u"4.3  Бусад богино хугацаат авлага (төрлөөр ангилна)"

    name = fields.Char(u'Үзүүлэлт')
    income = fields.Float(u'Эхний үлдэгдэл')
    expense = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage5(models.Model):
    _name = "account.disclosure.main.page5"
    _description = u"5. БУСАД САНХҮҮГИЙН ХӨРӨНГӨ"

    name = fields.Char(u'Төрөл')
    income = fields.Float(u'Эхний үлдэгдэл')
    expense = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage6(models.Model):
    _name = "account.disclosure.main.page6"
    _description = u"6. БАРАА МАТЕРИАЛ"

    name = fields.Char(u'Үзүүлэлт')
    raw_materials = fields.Float(u'Түүхий эд материал')
    mrp = fields.Float(u'Дуусаагүй үйлдвэрлэл')
    production = fields.Float(u'Бэлэн бүтээгдэхүүн')
    product = fields.Float(u'Бараа ')
    supply_materials = fields.Float(u'Хангамжийн материал')
    amount = fields.Float(u'Нийт дүн')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage8(models.Model):
    _name = "account.disclosure.main.page8"
    _description = u"8. УРЬДЧИЛЖ ТӨЛСӨН ЗАРДАЛ/ТООЦОО"

    name = fields.Char(u'Үзүүлэлт')
    first_amount = fields.Float(u'Эхний үлдэгдэл')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')
    table5_id = fields.Many2one('account.disclosure.report.table5')    


class AccountDisclosureReportMainPage9_1(models.Model):
    _name = "account.disclosure.main.page9_1"
    _description = u"9. ҮНДСЭН ХӨРӨНГӨ "

    name = fields.Char(u'Үзүүлэлт')
    gazriin_saijruulalt = fields.Char(u'Газрын сайжруулалт')
    barilga = fields.Char(u'Барилга, байгууламж')
    mashin_totu = fields.Char(u'Машин, тоног төхөөрөмж')
    teevriin = fields.Char(u'Тээврийн хэрэгсэл')
    tavilga = fields.Char(u'Тавилга эд хогшил')
    computer = fields.Char(u'Компьютер, бусад хэрэгсэл')
    busad = fields.Char(u'Бусад үндсэн хөрөнгө')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage9_2(models.Model):
    _name = "account.disclosure.main.page9_2"
    _description = u"9. ҮНДСЭН ХӨРӨНГӨ "

    name = fields.Char(u'Үзүүлэлт')
    gazriin_saijruulalt = fields.Char(u'Газрын сайжруулалт')
    barilga = fields.Char(u'Барилга, байгууламж')
    mashin_totu = fields.Char(u'Машин, тоног төхөөрөмж')
    teevriin = fields.Char(u'Тээврийн хэрэгсэл')
    tavilga = fields.Char(u'Тавилга эд хогшил')
    computer = fields.Char(u'Компьютер, бусад хэрэгсэл')
    busad = fields.Char(u'Бусад үндсэн хөрөнгө')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage10(models.Model):
    _name = "account.disclosure.main.page10"
    _description = u"10. БИЕТ БУС ХӨРӨНГӨ"

    name = fields.Char(u'Үзүүлэлт')
    zohiogchiin_erh = fields.Char(u'Зохиогчийн эрх')
    computer = fields.Char(u'Компьютерийн Програм хангамж')
    patent = fields.Char(u'Патент')
    baraanii_temdeg = fields.Char(u'Барааны тэмдэг')
    tusgai_zovshoorol = fields.Char(u'Тусгай зөвшөөрөл')
    gazar = fields.Char(u'Газар эзэмших эрх')
    biet_bus_horongo = fields.Char(u'Бусад биет бус хөрөнгө')
    total = fields.Char(u'Бүгд')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage11(models.Model):
    _name = "account.disclosure.main.page11"
    _description = u"11. ДУУСААГҮЙ БАРИЛГА"

    name = fields.Char(u'Дуусаагүй барилгын нэр')
    ehelsen_ognoo = fields.Date(u'Эхэлсэн он')
    duusgalin_huvi = fields.Float(u'Дуусгалтын хувь')
    niit_tosovt_ortog = fields.Float(u'Нийт төсөвт өртөг')
    etssiin_hugatsaa = fields.Date(u'Ашиглалтанд орох эцсийн хугацаа')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage12(models.Model):
    _name = "account.disclosure.main.page12"
    _description = u"12. БИОЛОГИЙН ХӨРӨНГӨ"

    name = fields.Char(u'Биологийн хөрөнгийн төрөл')
    age = fields.Float(u'Нас')
    gender = fields.Float(u'Хүйс')
    first_qty = fields.Float(u'Эхний үлдэгдэл - Тоо')
    first_amount = fields.Float(u'Эхний үлдэгдэл - Дансны үнэ')
    last_qty = fields.Float(u'Эцсийн үлдэгдэл - Тоо')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл - Дансны үнэ')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage13(models.Model):
    _name = "account.disclosure.main.page13"
    _description = u"13. УРТ ХУГАЦААТ ХӨРӨНГӨ ОРУУЛАЛТ"

    name = fields.Char(u'Хөрөнгө оруулалтын төрөл')
    first_percent = fields.Float(u'Эхний үлдэгдэл - Хөрөнгө оруулалтын хувь')
    first_amount = fields.Float(u'Эхний үлдэгдэл - Хөрөнгө оруулалтын дүн')
    last_percent = fields.Float(u'Эцсийн үлдэгдэл - Хөрөнгө оруулалтын хувь')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл - Хөрөнгө оруулалтын дүн')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage15(models.Model):
    _name = "account.disclosure.main.page15"
    _description = u"15. БУСАД ЭРГЭЛТИЙН БУС ХӨРӨНГӨ"

    name = fields.Char(u'Төрөл')
    first_amount = fields.Float(u'Эхний үлдэгдэл')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage16_1(models.Model):
    _name = "account.disclosure.main.page16_1"
    _description = u"16.1 Дансны өглөг"

    name = fields.Char(u'Ангилал')
    first_amount = fields.Float(u'Эхний үлдэгдэл')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage16_2(models.Model):
    _name = "account.disclosure.main.page16_2"
    _description = u"16.2 Татварын өр "

    name = fields.Char(u'Татварын өрийн төрөл')
    first_amount = fields.Float(u'Эхний үлдэгдэл')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage16_3(models.Model):
    _name = "account.disclosure.main.page16_3"
    _description = u"16.3 Богино хугацаат зээл"

    name = fields.Char(u'Үзүүлэлт')
    first_amount = fields.Float(u'Эхний үлдэгдэл - төгрөгөөр')
    first_amount_curr = fields.Float(u'Эхний үлдэгдэл - валютаар')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл - төгрөгөөр')
    last_amount_curr = fields.Float(u'Эцсийн үлдэгдэл - валютаар')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage16_4(models.Model):
    _name = "account.disclosure.main.page16_4"
    _description = u"16.4 Богино хугацаат нөөц (өр төлбөр)"

    name = fields.Char(u'Үзүүлэлт')
    first_amount = fields.Float(u'Эхний үлдэгдэл')
    add = fields.Float(u'Нэмэгдсэн **')
    sub = fields.Float(u'Хасагдсан (ашигласан нөөц)')
    useless = fields.Float(u'Ашиглаагүй буцаан бичсэн дүн')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage16_5(models.Model):
    _name = "account.disclosure.main.page16_5"
    _description = u"16.5 Бусад богино хугацаат өр төлбөр"

    name = fields.Char(u'Төрөл')
    first_amount = fields.Float(u'Эхний үлдэгдэл')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage16_6(models.Model):
    _name = "account.disclosure.main.page16_6"
    _description = u"16.6 Урт хугацаат зээл болон бусад урт хугацаат өр төлбөр"

    name = fields.Char(u'Үзүүлэлт')
    first_amount = fields.Float(u'Эхний үлдэгдэл - төгрөгөөр')
    first_amount_curr = fields.Float(u'Эхний үлдэгдэл - валютаар')
    last_amount = fields.Float(u'Эцсийн үлдэгдэл - төгрөгөөр')
    last_amount_curr = fields.Float(u'Эцсийн үлдэгдэл - валютаар')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage17_1(models.Model):
    _name = "account.disclosure.main.page17_1"
    _description = u"17.1 Өмч"

    name = fields.Char(u'Үзүүлэлт')
    qty_1 = fields.Float(u'Эргэлтэнд байгаа бүрэн төлөгдсөн энгийн хувьцаа - Тоо ширхэг')
    amount_1 = fields.Float(u'Эргэлтэнд байгаа бүрэн төлөгдсөн энгийн хувьцаа - Дүн (төгрөгөөр) ')
    qty_2 = fields.Float(u'Давуу эрхтэй хувьцаа - Тоо ширхэг')
    amount_2 = fields.Float(u'Давуу эрхтэй хувьцаа - Дүн (төгрөгөөр)')
    main_id = fields.Many2one('account.disclosure.main')


class AccountDisclosureReportMainPage17_2(models.Model):
    _name = "account.disclosure.main.page17_2"
    _description = u"17.2 Хөрөнгийн дахин үнэлгээний нэмэгдэл"

    name = fields.Char(u'Үзүүлэлт')
    undsen_hurungiin_dun = fields.Float(u'Үндсэн хөрөнгийн дахин үнэлгээний нэмэгдэл')
    biet_bus_hurungiin_dun = fields.Float(u'Биет бус  хөрөнгийн дахин үнэлгээний нэмэгдэл')
    main_id = fields.Many2one('account.disclosure.main')


# 17.3 Гадаад валютын хөрвүүлэлтийн нөөц
class AccountDisclosureReportMainPage17_3(models.Model):
    _name = "account.disclosure.main.page17_3"
    _description = u"17.3 Гадаад валютын хөрвүүлэлтийн нөөц"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', required=True)
    initial_balance = fields.Float(u'Эхний үлдэгдэл', required=True, default=0.0)
    nemegdsen = fields.Float(u'Нэмэгдсэн', required=True, default=0.0)
    hasagdsan = fields.Float(u'Хасагдсан', required=True, default=0.0)


# 18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ
class AccountDisclosureReportMainPage18_1(models.Model):
    _name = "account.disclosure.main.page18_1"
    _description = u"18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', required=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ
class AccountDisclosureReportMainPage18_2(models.Model):
    _name = "account.disclosure.main.page18_2"
    _description = u"18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', required=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ
class AccountDisclosureReportMainPage18_3(models.Model):
    _name = "account.disclosure.main.page18_3"
    _description = u"18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', required=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ
class AccountDisclosureReportMainPage18_4(models.Model):
    _name = "account.disclosure.main.page18_4"
    _description = u"18. БОРЛУУЛАЛТЫН ОРЛОГО БОЛОН БОРЛУУЛАЛТЫН ӨРТӨГ"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', required=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 19.1 Бусад орлого
class AccountDisclosureReportMainPage19_1(models.Model):
    _name = "account.disclosure.main.page19_1"
    _description = u"19.1 Бусад орлого"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Орлогын төрөл', required=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 19.2 Гадаад валютын ханшийн зөрүүний олз, гарз
class AccountDisclosureReportMainPage19_2(models.Model):
    _name = "account.disclosure.main.page19_2"
    _description = u"19.2 Гадаад валютын ханшийн зөрүүний олз, гарз"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Төрөл', readonly=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 19.3 Бусад ашиг (алдагдал)
class AccountDisclosureReportMainPage19_3(models.Model):
    _name = "account.disclosure.main.page19_3"
    _description = u"19.3 Бусад ашиг (алдагдал)"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Ашиг (алдагдал)')
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 20.1 Борлуулалт маркетингийн болон ерөнхий ба удирдлагын зардлууд
class AccountDisclosureReportMainPage20_1(models.Model):
    _name = "account.disclosure.main.page20_1"
    _description = u"20.1 Борлуулалт маркетингийн болон ерөнхий ба удирдлагын зардлууд"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Зардлын төрөл')
    last_year_sale_amount = fields.Float(u'Өмнөх оны дүн: БорМар', required=True, default=0.0)
    last_year_management_amount = fields.Float(u'Өмнөх оны дүн: ЕрУд', required=True, default=0.0)
    this_year_sale_amount = fields.Float(u'Тайлант оны дүн: БорМар', required=True, default=0.0)
    this_year_management_amount = fields.Float(u'Тайлант оны дүн: ЕрУд', required=True, default=0.0)


# 20.2 Бусад зардал
class AccountDisclosureReportMainPage20_2(models.Model):
    _name = "account.disclosure.main.page20_2"
    _description = u"20.2 Бусад зардал"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Зардлын төрөл', readonly=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 20.3 Цалингийн зардал
class AccountDisclosureReportMainPage20_3(models.Model):
    _name = "account.disclosure.main.page20_3"
    _description = u"20.3 Цалингийн зардал"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Ангилал', readonly=True)
    avarage_employees = fields.Integer(u'Ажиллагчдын дундаж тоо', required=True, default=0)
    last_year_amount = fields.Float(u'Цалингийн зардлын дүн: Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Цалингийн зардлын дүн: Тайлант оны дүн', required=True, default=0.0)


# 21. ОРЛОГЫН ТАТВАРЫН ЗАРДАЛ
class AccountDisclosureReportMainPage21(models.Model):
    _name = "account.disclosure.main.page21"
    _description = u"21. ОРЛОГЫН ТАТВАРЫН ЗАРДАЛ"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', readonly=True)
    last_year_amount = fields.Float(u'Өмнөх оны дүн', required=True, default=0.0)
    this_year_amount = fields.Float(u'Тайлант оны дүн', required=True, default=0.0)


# 22.1 Толгой компани, хамгийн дээд хяналт тавигч компани, хувь хүний талаарх мэдээлэл****
class AccountDisclosureReportMainPage22_1(models.Model):
    _name = "account.disclosure.main.page22_1"
    _description = u"22.1 Толгой компани, хамгийн дээд хяналт тавигч компани, хувь хүний талаарх мэдээлэл****"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', readonly=True)
    head_company = fields.Char(u'Толгой компани')
    hamgiin_deed_hyanalt_tavigch_tolgoi_company = fields.Char(u'Хамгийн дээд хяналт тавигч толгой компани')
    hamgiin_deed_hyanalt_tavigch_huvi_hun = fields.Char(u'Хамгийн дээд хяналт тавигч хувь хүн')
    description = fields.Char(u'Тайлбар')


# 22.2 Тэргүүлэх удирдлагын бүрэлдэхүүнд олгосон нөхөн олговрын тухай мэдээлэл
class AccountDisclosureReportMainPage22_2(models.Model):
    _name = "account.disclosure.main.page22_2"
    _description = u"22.2 Тэргүүлэх удирдлагын бүрэлдэхүүнд олгосон нөхөн олговрын тухай мэдээлэл"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', readonly=True)
    initial_balance = fields.Float(u'Эхний үлдэгдэл', required=True, default=0.0)
    end_balance = fields.Float(u'Эцсийн үлдэгдэл', required=True, default=0.0)


# 22.3 Холбоотой талуудтай хийсэн ажил гүйлгээ
class AccountDisclosureReportMainPage22_3(models.Model):
    _name = "account.disclosure.main.page22_3"
    _description = u"22.3 Холбоотой талуудтай хийсэн ажил гүйлгээ"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Холбоотой талын нэр', required=True)
    transaction_description = fields.Char(u'Ажил гүйлгээний утга', required=True)
    amount = fields.Float(u'Дүн', required=True, default=0.0)
    description = fields.Char(u'Тайлбар', required=True)


# 25. ХӨРӨНГӨ ОРУУЛАЛТ
class AccountDisclosureReportMainPage25(models.Model):
    _name = "account.disclosure.main.page25"
    _description = u"25. ХӨРӨНГӨ ОРУУЛАЛТ"

    main_id = fields.Many2one('account.disclosure.main')

    name = fields.Char(u'Үзүүлэлт', readonly=True)
    initial_balance = fields.Float(u'Эхний үлдэгдэл', required=True, default=0.0)
    aj_ahui_negjiin_ooriin_horongoor = fields.Float(u'Аж ахуйн нэгжийн өөрийн хөрөнгөөр', required=True, default=0.0)
    ulsiin_tosviin_horongoor = fields.Float(u'Улсын төсвийн хөрөнгөөр', required=True, default=0.0)
    oron_nutgiin_tosviin_horongoor = fields.Float(u'Орон нутгийн төсвийн хөрөнгөөр', required=True, default=0.0)
    banknii_zeel = fields.Float(u'Банкны зээл', required=True, default=0.0)
    gadaadiin_shuud_horongo_oruulalt = fields.Float(u'Гадаадын шууд хөрөнгө оруулалт', required=True, default=0.0)
    gadaadiin_zeel = fields.Float(u'Гадаадын зээл', required=True, default=0.0)
    gadaadiin_butsaltgui_tuslamj = fields.Float(u'Гадаадын буцалтгүй тусламж', required=True, default=0.0)
    tosol_hotolbor_handiv = fields.Float(u'Төсөл хөтөлбөр хандив', required=True, default=0.0)
    busad_eh_uusver = fields.Float(u'Бусад эх үүсвэр', required=True, default=0.0)
